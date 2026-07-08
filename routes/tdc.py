from flask import Blueprint, request, jsonify, current_app
import db
import json
import os
from werkzeug.utils import secure_filename
import time
import re
from auth.decorator import login_required, permission_required
tdc_bp = Blueprint('tdc_bp', __name__)

# Config folder upload
UPLOAD_FOLDER = 'static/uploads/pdfs'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@tdc_bp.route('/tdc_manager', methods=['GET'])
@permission_required('tdc_view')
def tdc_manager_page():
    # Render file html mới
    return current_app.jinja_env.get_template('tdc_manager.html').render()

# API 1: Lấy danh sách thư viện TDC (Chỉ Active)
@tdc_bp.route('/api/get_tdc_library', methods=['GET'])
@permission_required('tdc_view')
def get_tdc_library():
    # Sử dụng hàm mới lấy list Active
    data = db.get_active_tdc_list()
    for d in data:
        if d['criteria_json']: d['criteria'] = json.loads(d['criteria_json'])
    return jsonify(data)

# [MỚI] API 1b: Lấy danh sách Pending (Chờ duyệt)
@tdc_bp.route('/api/get_tdc_pending', methods=['GET'])
@permission_required('tdc_approval')
def get_tdc_pending():
    data = db.get_tdc_pending_list()
    for d in data:
        if d['criteria_json']: d['criteria'] = json.loads(d['criteria_json'])
    return jsonify(data)

# [MỚI] API 1c: Lấy lịch sử Version
@tdc_bp.route('/api/get_tdc_history', methods=['POST'])
@permission_required('tdc_view')
def get_tdc_history():
    req = request.json
    master_id = req.get('master_id')
    if not master_id: return jsonify({'status':'error', 'msg':'Thiếu Master ID'})
    
    data = db.get_tdc_history(master_id)
    for d in data:
        if d['criteria_json']: d['criteria'] = json.loads(d['criteria_json'])
    return jsonify({'status':'success', 'data': data})

# API 2: Upload PDF
@tdc_bp.route('/api/upload_tdc_pdf', methods=['POST'])
def upload_tdc_pdf():
    if 'file' not in request.files: return jsonify({'status':'error', 'msg':'No file'})
    f = request.files['file']
    if f.filename == '': return jsonify({'status':'error', 'msg':'No filename'})
    
    if f:
        filename = secure_filename(f"{int(time.time())}_{f.filename}")
        path = os.path.join(UPLOAD_FOLDER, filename)
        f.save(path)
        return jsonify({'status':'success', 'path': f"/{path}"}) # Trả về đường dẫn web

# API 3: Lưu TDC Version (Draft / Pending)
@tdc_bp.route('/api/save_tdc_master', methods=['POST'])
@permission_required('tdc_editor')
def save_tdc_master_endpoint():
    req = request.json
    conn = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        # Dữ liệu từ form
        code = req.get('tdc_code')
        cust = req.get('customer_name')
        purpose = req.get('usage_purpose')
        grade = req.get('grade')
        
        # Validation cơ bản phía server
        if not code or not cust:
            return jsonify({'status': 'error', 'msg': 'Thiếu thông tin Code hoặc Khách hàng'})

        # --- BƯỚC 1: XỬ LÝ MASTER (Tìm hoặc Tạo) ---
        cursor.execute("SELECT id FROM tdc_master WHERE tdc_code = ? AND is_deleted = 0", (code,))
        row = cursor.fetchone()
        
        master_id = None
        
        if row:
            # A. Đã tồn tại -> Lấy ID cũ
            master_id = row[0]
            cursor.execute("""
                UPDATE tdc_master 
                SET customer_name = ?, usage_purpose = ?, grade = ? 
                WHERE id = ?
            """, (cust, purpose, grade, master_id))
        else:
            # B. Chưa có -> Tạo mới (FIX LỖI NO RESULTS)
            # Thêm "SET NOCOUNT ON;" để tắt thông báo "1 row affected" gây nhiễu
            cursor.execute("""
                SET NOCOUNT ON;
                INSERT INTO tdc_master (tdc_code, customer_name, usage_purpose, grade, created_at,is_deleted)
                VALUES (?, ?, ?, ?, GETDATE(),0);
                SELECT CAST(SCOPE_IDENTITY() AS INT);
            """, (code, cust, purpose, grade))
            result = cursor.fetchone()
            if result and result[0]:
                master_id = result[0]
            else:
                cursor.execute("SELECT id FROM tdc_master WHERE tdc_code = ? AND is_deleted = 0", (code,))
                res_check = cursor.fetchone()
                if res_check:
                    master_id = res_check[0]
                else:
                    raise Exception("Lỗi hệ thống: Không lấy được ID sau khi INSERT.")

        # --- BƯỚC 2: TÍNH VERSION TIẾP THEO ---
        # Logic: Nếu chưa có version nào (bảng rỗng hoặc master mới) -> ver 1. Nếu có rồi -> max + 1
        cursor.execute("SELECT ISNULL(MAX(version_no), 0) + 1 FROM tdc_versions WHERE master_id = ?", (master_id,))
        next_ver = cursor.fetchone()[0]

        # --- BƯỚC 3: TẠO VERSION MỚI (Pending) ---
        cursor.execute("""
            INSERT INTO tdc_versions 
            (master_id, version_no, criteria_json, pdf_path, valid_from, valid_to, status, created_at, note_1, note_2)
            VALUES (?, ?, ?, ?, ?, ?, 'Pending', GETDATE(), ?, ?)
        """, (
            master_id, 
            next_ver, 
            json.dumps(req.get('criteria', [])), 
            req.get('pdf_path', ''),
            req.get('valid_from') or None, 
            req.get('valid_to') or None,
            req.get('note_1', ''), 
            req.get('note_2', '')  
        ))
        
        conn.commit()
        return jsonify({'status': 'success', 'msg': f'Đã lưu {code} - Phiên bản v{next_ver} (Chờ duyệt)'})

    except Exception as e:
        if conn: conn.rollback()
        print("Lỗi Save TDC:", str(e))
        return jsonify({'status': 'error', 'msg': f"Lỗi Server: {str(e)}"})
    finally:
        if conn: conn.close()

@tdc_bp.route('/api/confirm_tdc', methods=['POST'])
@permission_required('tdc_approval')
def confirm_tdc_endpoint():
    req = request.json
    # Tiếp nhận dữ liệu
    data = {
    'version_id': req.get('version_id'),
    'cust': req.get('customer_name'),
    'grade': req.get('grade'),
    'purpose': req.get('usage_purpose'),
    'tdc_code': req.get('tdc_code'), 
    'valid_from': req.get('valid_from'),
    'valid_to': req.get('valid_to'),
    'criteria': req.get('criteria', []),
    'note_1': req.get('note_1', ''), 
    'note_2': req.get('note_2', '')
}
    
    # --- [VALIDATION QUAN TRỌNG] ---
    if not data['version_id']: 
        return jsonify({'status':'error', 'msg':'Lỗi: Thiếu Version ID'})
    
    # Thêm đoạn này để chặn lỗi tên khách hàng rỗng
    if not data['cust'] or not str(data['cust']).strip():
        return jsonify({'status':'error', 'msg':'Lỗi: Tên Khách Hàng không được để trống khi Duyệt!'})

    if not data['grade'] or not str(data['grade']).strip():
        return jsonify({'status':'error', 'msg':'Lỗi: Mác thép không được để trống!'})
    conn = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # 1. Lấy Master ID
        cursor.execute("SELECT master_id FROM tdc_versions WHERE id = ?", (data['version_id'],))
        row = cursor.fetchone()
        if not row: return jsonify({'status':'error', 'msg':'Không tìm thấy Version'})
        master_id = row[0]
        
        # 2. Kiểm tra Overlap (Trùng lặp thời gian)
        # Bỏ qua chính master_id này để cho phép Edit/Gia hạn
        cursor.execute("""
            SELECT v.id, m.tdc_code 
            FROM tdc_versions v
            JOIN tdc_master m ON v.master_id = m.id
            WHERE v.status = 'Active'
            AND v.master_id != ?
            AND m.customer_name = ?
            AND m.usage_purpose = ?
            AND m.tdc_code = ?
        """, (master_id, data['cust'], data['purpose'], data['tdc_code']))

        conflict = cursor.fetchone()
        if conflict:
            return jsonify({'status': 'error', 'msg': f"Lỗi: Đã có TDC cho khách hàng này với cùng Mục đích và Mã TDC đang ở trạng thái Active!"})

        # 3. THỰC HIỆN UPDATE & CONFIRM
        # A. Chuyển bản Active cũ sang Replaced
        cursor.execute("UPDATE tdc_versions SET status = 'Replaced' WHERE master_id = ? AND status = 'Active'", (master_id,))
        
        # B. Kích hoạt bản mới
        cursor.execute("""
            UPDATE tdc_versions 
            SET status = 'Active', 
                valid_from = ?, 
                valid_to = ?, 
                confirmed_at = GETDATE(), 
                note_1 = ?, 
                note_2 = ?
            WHERE id = ?
        """, (
            data.get('valid_from') if data.get('valid_from') else None, 
            data.get('valid_to') if data.get('valid_to') else None, 
            data['note_1'], 
            data['note_2'], 
            data['version_id']
        ))
        
        # C. [QUAN TRỌNG] Cập nhật lại Master Name
        cursor.execute("""
            UPDATE tdc_master
            SET customer_name = ?, grade = ?, usage_purpose = ?
            WHERE id = ?
        """, (data['cust'], data['grade'], data['purpose'], master_id))

        conn.commit()
        return jsonify({'status':'success', 'msg': 'Đã duyệt thành công!'})

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'status': 'error', 'msg': f"Lỗi hệ thống: {str(e)}"})
    finally:
        if conn: conn.close()

# API: Thay đổi trạng thái sang Rejected
@tdc_bp.route('/api/reject_tdc', methods=['POST'])
@permission_required('tdc_approval')
def reject_tdc_endpoint():
    req = request.json
    ver_id = req.get('version_id')
    reason = req.get('reason', '')
    if not ver_id: return jsonify({'status':'error', 'msg':'Missing Version ID'})
    
    success, msg = db.reject_tdc_version(ver_id, reason)
    if success:
        return jsonify({'status':'success', 'msg': msg})
    else:
        return jsonify({'status':'error', 'msg': msg})

@tdc_bp.route('/api/sap/get_customers', methods=['GET'])
def get_sap_customers():
    conn = None
    try:
        conn = db.get_connection()
        # Lấy danh sách khách hàng duy nhất từ bảng SO
        query = "SELECT DISTINCT [Customer] FROM [factory].[dbo].[so] WITH (NOLOCK) ORDER BY [Customer]"
        cursor = conn.cursor()
        cursor.execute(query)
        rows = cursor.fetchall() # Trả về list of tuples      
        # Chuyển về list string đơn giản
        customers = [row[0] for row in rows if row[0]]
        return jsonify({'status': 'success', 'data': customers})
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn:
            conn.close()
# --- [MỚI] API 6: Lấy Mác thép theo Khách hàng từ SAP ---
@tdc_bp.route('/api/sap/get_grades_by_cust', methods=['POST'])
def get_grades_by_cust():
    conn = None
    try:
        req = request.json
        cust_name = req.get('customer_name')
        if not cust_name: return jsonify({'status':'error', 'msg':'Thiếu tên khách hàng'})

        conn = db.get_connection()
        # Lấy Description của khách hàng này để parse Mác thép
        query = "SELECT DISTINCT [Item Description] FROM [factory].[dbo].[so] WITH (NOLOCK) WHERE [Customer] = ?"
        cursor = conn.cursor()
        cursor.execute(query, (cust_name,))
        rows = cursor.fetchall()

        # Logic Parse Mác thép (Giống bên Allocation)
        grades = set()
        for r in rows:
            desc = r[0] # VD: Thép HRC HSPM 2.00x1250 SPHC
            if not desc: continue
            
            # Regex tìm kích thước (VD: 2.00x1250)
            dim_match = re.search(r'(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)', desc)
            if dim_match:
                # Lấy phần chuỗi SAU kích thước
                remaining = desc[dim_match.end():].strip()
                if remaining:
                    # Mác thép thường là từ đầu tiên sau kích thước
                    parts = remaining.split()
                    raw_grade = parts[0].strip().upper()
                    # Loại bỏ các từ rác nếu cần (VD: COIL, PICKLED...)
                    if len(raw_grade) > 2: # Mác thép ít nhất 3 ký tự
                        grades.add(raw_grade)

        return jsonify({'status': 'success', 'data': sorted(list(grades))})
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn:
            conn.close()
@tdc_bp.route('/api/activate_specific_version', methods=['POST'])
@permission_required('tdc_editor')
def activate_specific_version():
    req = request.json
    ver_id = req.get('version_id')
    master_id = req.get('master_id')
    
    if not ver_id or not master_id:
        return jsonify({'status': 'error', 'msg': 'Thiếu ID phiên bản'})
    conn = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # 1. Đưa tất cả các phiên bản của Master này về trạng thái Replaced
        cursor.execute("UPDATE tdc_versions SET status = 'Replaced' WHERE master_id = ?", (master_id,))
        
        # 2. Kích hoạt đúng phiên bản được chọn
        cursor.execute("UPDATE tdc_versions SET status = 'Active' WHERE id = ?", (ver_id,))
        
        conn.commit()
        return jsonify({'status': 'success', 'msg': 'Đã chuyển trạng thái Active thành công!'})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
@tdc_bp.route('/api/get_tdc_complaints', methods=['POST'])
def get_tdc_complaints():
    req = request.json
    master_id = req.get('master_id')
    if not master_id: return jsonify({'status':'error', 'msg':'Thiếu Master ID'})
    
    # Lấy data từ DB (đã có sẵn c.version_id JOIN ra v.version_no)
    data = db.get_complaints_by_master(master_id) 
    
    formatted_data = []
    for item in data:
        formatted_data.append({
            'id': item['id'], 
            'date': item['complaint_date'].strftime('%d/%m/%Y'),
            'group': item['defect_group'],
            'desc': item['description'],
            'resolution': item['resolution_plan'] if item['resolution_plan'] else "", 
            'weight': f"{item['weight_kg']:,} kg" if item['weight_kg'] else "-",
            'amount': f"{item['amount_vnd']:,} VNĐ" if item['amount_vnd'] else "-",
            'status': item['status'],
            
            # --- [THÊM DÒNG NÀY ĐỂ GIAO DIỆN HẾT BỊ UNDEFINED] ---
            # Dùng .get() với giá trị mặc định '?' để phòng ngừa các khiếu nại cũ bị Null
            'version': item.get('version_no', '?'), 
            
            'so_number': item['so_number'] if item['so_number'] else "",
            'coil_ids': item['coil_ids'] if item['coil_ids'] else "",
            'pdf_path': item['pdf_path'] if item['pdf_path'] else ""
        })
    return jsonify({'status': 'success', 'data': formatted_data})
@tdc_bp.route('/api/update_complaint_status', methods=['POST'])
def update_complaint_status_endpoint():
    req = request.json
    comp_id = req.get('id')
    new_status = req.get('status')
    
    if not comp_id or not new_status:
        return jsonify({'status':'error', 'msg':'Thiếu thông tin'})
        
    success, msg = db.update_complaint_status(comp_id, new_status)
    if success:
        return jsonify({'status':'success', 'msg': msg})
    else:
        return jsonify({'status':'error', 'msg': msg})
@tdc_bp.route('/api/edit_complaint_full', methods=['POST'])
def edit_complaint_full():
    req = request.json
    # Validate cơ bản
    if not req.get('id'):
        return jsonify({'status':'error', 'msg':'Thiếu ID khiếu nại'})
    
    success, msg = db.update_complaint_details(req)
    
    if success:
        return jsonify({'status':'success', 'msg': msg})
    else:
        return jsonify({'status':'error', 'msg': msg})
@tdc_bp.route('/api/add_tdc_complaint', methods=['POST'])
def add_tdc_complaint():
    req = request.json
    # Gọi hàm từ db.py cho sạch code
    success, msg = db.add_complaint(req)
    
    if success:
        return jsonify({'status': 'success', 'msg': msg})
    else:
        return jsonify({'status': 'error', 'msg': msg})
COMPLAINT_UPLOAD_FOLDER = 'static/uploads/complaints'
os.makedirs(COMPLAINT_UPLOAD_FOLDER, exist_ok=True)
# API Upload cho Khiếu nại
@tdc_bp.route('/api/upload_complaint_pdf', methods=['POST'])
def upload_complaint_pdf():
    # Sử dụng getlist('files') thay vì files['file']
    files = request.files.getlist('files')
    if not files or len(files) == 0:
        return jsonify({'status':'error', 'msg':'No files uploaded'})
    
    uploaded_paths = []
    for f in files:
        if f.filename != '':
            filename = secure_filename(f"comp_{int(time.time())}_{f.filename}")
            path = os.path.join(COMPLAINT_UPLOAD_FOLDER, filename)
            f.save(path)
            uploaded_paths.append(f"/{path}")
            
    if uploaded_paths:
        return jsonify({'status':'success', 'paths': uploaded_paths}) # Trả về mảng paths
    else:
        return jsonify({'status':'error', 'msg':'Lỗi không lưu được file nào'}) 
SECRET_DELETE_CODE = "ADMIN@123"
@tdc_bp.route('/api/delete_tdc', methods=['POST'])
@permission_required('tdc_editor')
def delete_tdc_endpoint():
    req = request.json
    master_id = req.get('id')
    auth_code = req.get('auth_code')
    if not master_id:
        return jsonify({'status': 'error', 'msg': 'Thiếu ID của TDC cần xóa'})
    if auth_code != SECRET_DELETE_CODE:
        return jsonify({'status': 'error', 'msg': 'Mã xác nhận không chính xác! Giao dịch bị từ chối.'})   
    success, msg = db.delete_tdc(master_id)
    if success:
        return jsonify({'status':'success', 'msg': msg})
    else:
        return jsonify({'status':'error', 'msg': msg})
@tdc_bp.route('/tdc_dashboard', methods=['GET'])
def tdc_dashboard_page():
    return current_app.jinja_env.get_template('complaint_dashboard.html').render()

@tdc_bp.route('/api/dashboard_complaints', methods=['POST'])
def api_dashboard_complaints():
    req = request.json
    month_filter = req.get('month') # Định dạng YYYY-MM
    data = db.get_dashboard_stats(month_filter)
    return jsonify({'status': 'success', 'data': data})
import io
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file, request, jsonify
import db

@tdc_bp.route('/api/export_tdc_excel', methods=['POST'])
def export_tdc_excel():
    req = request.json
    selected_codes = req.get('tdc_codes', [])
    
    if not selected_codes:
        return jsonify({'status': 'error', 'msg': 'Không có TDC Code nào được chọn'}), 400

    # 1. LẤY DỮ LIỆU TỪ DB
    all_tdcs = db.get_active_tdc_list()
    filtered_tdcs = [t for t in all_tdcs if t.get('tdc_code') in selected_codes]

    history_map = {}
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Tạo chuỗi tham số an toàn (?,?,?)
        placeholders = ','.join(['?'] * len(selected_codes))
        
        # JOIN hai bảng qua master_id và gọi chính xác tên cột version_no
        cursor.execute(f"""
            SELECT m.tdc_code, v.version_no, v.created_at 
            FROM tdc_versions v
            INNER JOIN tdc_master m ON v.master_id = m.id
            WHERE m.tdc_code IN ({placeholders})
            ORDER BY m.tdc_code, v.version_no ASC
        """, tuple(selected_codes))
        
        for row in cursor.fetchall():
            t_code = row[0]
            t_ver = row[1]   # Chính là version_no (1, 2, 3...)
            t_date = row[2]  # Chính là created_at
            
            if t_code not in history_map:
                history_map[t_code] = {}
            
            # Format ngày tháng ra dạng YYYY-MM-DD HH:MM
            date_str = ""
            if t_date:
                try:
                    date_str = t_date.strftime("%Y-%m-%d %H:%M")
                except:
                    date_str = str(t_date)
                    
            # Ghi vào dictionary với key dạng V1, V2, V3...
            history_map[t_code][f"V{t_ver}"] = date_str
    except Exception as e:
        print("Lỗi truy vấn lịch sử TDC:", e)

    # 2. ĐỊNH NGHĨA CỘT BẰNG LIST BẠN CUNG CẤP
    DEFECT_OPTS = [
        # --- BỀ MẶT (surf) ---
        { "code": "MI", "name": "TCPK nặng", "grp": "surf" },
        { "code": "HPrScale", "name": "Xỉ sơ cấp HP", "grp": "surf" },
        { "code": "EL", "name": "Lỗi xếp lớp", "grp": "surf" },
        { "code": "HOLE", "name": "Lỗ thủng", "grp": "surf" },
        { "code": "RIP", "name": "Rách bề mặt", "grp": "surf" },
        { "code": "BRUS", "name": "Vết Hằn trục", "grp": "surf" },
        { "code": "LC", "name": "Nứt dọc", "grp": "surf" },
        { "code": "SCRT", "name": "Xước bề mặt", "grp": "surf" },
        { "code": "TCPK-n", "name": "TCPK nhẹ", "grp": "surf" },
        { "code": "oil", "name": "Gấp nếp", "grp": "surf" },
        { "code": "rust", "name": "Nếp nhăn", "grp": "surf" },
        { "code": "scratch_m", "name": "Vết hằn Pinch Roll", "grp": "surf" },
        { "code": "dirt", "name": "Gãy mặt", "grp": "surf" },
        { "code": "XTC", "name": "Xỉ thứ cấp", "grp": "surf" },
        { "code": "XC", "name": "Xỉ cán", "grp": "surf" },
        { "code": "other_s", "name": "Xỉ muối tiêu", "grp": "surf" },
        { "code": "gianbien", "name": "Giãn biên/Bụng", "grp": "surf" },
        { "code": "chambi", "name": "Chấm bi", "grp": "surf" }, 
        # --- HÌNH HỌC (geo) ---
        { "code": "Crown", "name": "Độ Crown", "grp": "geo" },
        { "code": "Wedge", "name": "Độ Wedge", "grp": "geo" },
        { "code": "ThickDiff", "name": "Sai lệch dày", "grp": "geo" },
        { "code": "WidthDiff", "name": "Sai lệch rộng", "grp": "geo" },
        { "code": "telescope", "name": "Cong cạnh", "grp": "geo" },
        { "code": "high_spot", "name": "High Spot", "grp": "geo" },
        { "code": "dungsaitrong", "name": "Dung sai ĐK trong", "grp": "geo" },
        # --- NGOẠI QUAN (app) ---
        { "code": "strap", "name": "Khuyết biên", "grp": "app" },
        { "code": "label_tag", "name": "Bavia biên", "grp": "app" },
        { "code": "packaging", "name": "Vỡ biên", "grp": "app" },
        { "code": "edge_cond", "name": "Sổ vòng", "grp": "app" },
        { "code": "coil_shape", "name": "Loa cuộn", "grp": "app" },
        # --- CƠ LÝ HÓA (prop) ---
        { "code": "YieldPoint", "name": "GH Chảy", "grp": "prop" },
        { "code": "Tensile", "name": "GH Bền", "grp": "prop" },
        { "code": "Elongation", "name": "Độ giãn dài", "grp": "prop" },
        { "code": "Hardness", "name": "Độ cứng", "grp": "prop" },
        { "code": "ImpactEnergy", "name": "Độ dai va đập", "grp": "prop" },
        { "code": "C", "name": "Carbon", "grp": "prop" },
        { "code": "Mn", "name": "Mangan", "grp": "prop" },
        { "code": "Si", "name": "Silic", "grp": "prop" },
        { "code": "P", "name": "Photpho", "grp": "prop" },
        { "code": "S", "name": "Lưu huỳnh", "grp": "prop" },
        { "code": "Cu", "name": "Đồng", "grp": "prop" },
        { "code": "Ni", "name": "Niken", "grp": "prop" },
        { "code": "Cr", "name": "Crom", "grp": "prop" },
        { "code": "Mo", "name": "Moly", "grp": "prop" },
        { "code": "V", "name": "Vanadi", "grp": "prop" },
        { "code": "Ti", "name": "Titan", "grp": "prop" },
        { "code": "Al", "name": "Nhôm", "grp": "prop" },
        { "code": "Ca", "name": "Canxi", "grp": "prop" },
        { "code": "B", "name": "Bo", "grp": "prop" },
        { "code": "Nb", "name": "Niobi", "grp": "prop" },
        { "code": "CEV", "name": "CEV", "grp": "prop" },
        { "code": "O", "name": "Oxy", "grp": "prop" },
        { "code": "N", "name": "Nitơ", "grp": "prop" },
        { "code": "H", "name": "Hydro", "grp": "prop" },
    ]

    # [MỚI] Tự động sinh 15 cột Lịch sử cập nhật
    HISTORY_OPTS = [{"code": f"V{i}", "name": f"V{i}", "grp": "history"} for i in range(1, 16)]

    # Cột chung ưu tiên lên đầu
    COLUMNS = [
        {"code": "customer_name", "name": "Tên Khách hàng", "grp": "CHUNG"},
        {"code": "grade", "name": "Mác thép", "grp": "CHUNG"},
        {"code": "usage_purpose", "name": "Ứng dụng", "grp": "CHUNG"},
        {"code": "tdc_code", "name": "Mã TDC", "grp": "CHUNG"},
    ] + DEFECT_OPTS + HISTORY_OPTS

    # [MỚI] BẢNG MÀU CHUYÊN BIỆT CHO HEADER
    GROUP_CONFIG = {
        "CHUNG": {"name": "THÔNG TIN CHUNG", "bg": "E2E8F0", "text": "0F172A"},     # Xám
        "surf": {"name": "BỀ MẶT", "bg": "FEF08A", "text": "854D0E"},               # Vàng
        "geo": {"name": "KÍCH THƯỚC", "bg": "BAE6FD", "text": "0369A1"},            # Xanh da trời
        "app": {"name": "NGOẠI QUAN", "bg": "D9F99D", "text": "3F6212"},            # Xanh nhạt
        "prop": {"name": "CƠ LÝ HÓA", "bg": "166534", "text": "FFFFFF"},            # Xanh lá ĐẬM (chữ Trắng)
        "history": {"name": "LỊCH SỬ CẬP NHẬT", "bg": "FED7AA", "text": "9A3412"}   # Cam nhạt
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TDC_Data"

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 3. VẼ HEADER KÈM MÀU SẮC
    for col_idx, col_data in enumerate(COLUMNS, start=1):
        grp_key = col_data["grp"]
        cfg = GROUP_CONFIG.get(grp_key, {"name": grp_key, "bg": "FFFFFF", "text": "000000"})
        
        # Style màu nền và màu chữ
        fill_style = PatternFill(start_color=cfg["bg"], end_color=cfg["bg"], fill_type="solid")
        font_style = Font(bold=True, color=cfg["text"])
        
        c1 = ws.cell(row=1, column=col_idx, value=cfg["name"])
        c2 = ws.cell(row=2, column=col_idx, value=col_data["name"])
        c3 = ws.cell(row=3, column=col_idx, value=col_data["code"])
        
        for cell in [c1, c2, c3]:
            cell.font = font_style
            cell.alignment = center_align
            cell.fill = fill_style
            cell.border = thin_border

    # GỘP Ô THÔNG MINH CHO DÒNG 1 (Nhóm)
    start_col = 1
    current_grp = COLUMNS[0]["grp"]
    for col_idx in range(2, len(COLUMNS) + 2):
        grp = COLUMNS[col_idx-1]["grp"] if col_idx <= len(COLUMNS) else None
        
        if grp != current_grp:
            if col_idx - 1 > start_col:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col_idx-1)
            start_col = col_idx
            current_grp = grp

    # Ẩn dòng số 3
    ws.row_dimensions[3].hidden = True

    # 4. ĐỔ DỮ LIỆU TỪ DÒNG SỐ 4
    current_row = 4
    for tdc in filtered_tdcs:
        t_code = tdc.get('tdc_code', '')
        
        criteria_dict = {}
        if tdc.get('criteria_json'):
            try:
                criteria_list = json.loads(tdc['criteria_json'])
                for item in criteria_list:
                    defect = item.get('defect')
                    rng = item.get('range', [])
                    if defect and rng:
                        min_val, max_val = min(rng), max(rng)
                        # [MỚI] LOẠI BỎ IF-ELSE, LUÔN LUÔN THÊM "C" CHO TẤT CẢ CÁC NHÓM
                        criteria_dict[defect] = f"C{min_val}" if min_val == max_val else f"C{min_val} - C{max_val}"
            except:
                pass

        # Lấy mảng version đã truy vấn ở Bước 1.5
        tdc_hist = history_map.get(t_code, {})

        for col_idx, col_data in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            grp = col_data["grp"]
            code = col_data["code"]

            if grp == "CHUNG":
                cell.value = tdc.get(code, '')
            elif grp == "history":
                # Điền ngày cập nhật tương ứng với V1, V2... (nếu có)
                cell.value = tdc_hist.get(code, '') 
            else:
                cell.value = criteria_dict.get(code, '')
                
        current_row += 1

    # 5. CĂN CHỈNH ĐỘ RỘNG CỘT TỰ ĐỘNG
    for col_idx in range(1, len(COLUMNS) + 1):
        column_letter = get_column_letter(col_idx)
        header_val = ws.cell(row=2, column=col_idx).value
        header_len = len(str(header_val)) if header_val else 10
        ws.column_dimensions[column_letter].width = max(header_len + 4, 12)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(out, as_attachment=True, download_name="TDC_Export_Data.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')