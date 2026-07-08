from flask import Blueprint, render_template, request, jsonify, send_file, session
import json
import db  # Module db.py
import threading
import re
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from excel_processor import process_sales_order_excel, sync_excel_to_database
from auth.decorator import login_required, permission_required
alloc_run_bp = Blueprint('alloc_run_bp', __name__)
@alloc_run_bp.route('/api/upload_excel_so', methods=['POST'])
def upload_excel_so():
    conn = None # Khai báo biến conn ở ngoài cùng
    try:
        # 1. Nhận file từ Request
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'msg': 'Không tìm thấy file.'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'msg': 'Chưa chọn file.'})

        # 2. XỬ LÝ PANDAS (Hoàn toàn chạy trên RAM, KHÔNG MỞ DATABASE)
        # Hàm này chứa thuật toán regex tách CW, dọn rác, chuẩn hóa chữ...
        cleaned_records = process_sales_order_excel(file)
        
        if not cleaned_records:
            return jsonify({'status': 'error', 'msg': 'File Excel không có dữ liệu hợp lệ.'})

        # 3. MỞ KẾT NỐI DATABASE (Chỉ mở khi chuẩn bị Ghi)
        conn = db.get_connection()
        
        # 4. GỌI HÀM ĐỒNG BỘ (UPSERT)
        # Hàm sync_excel_to_database sẽ dùng conn này để chạy SQL và gọi conn.commit()
        inserted, updated = sync_excel_to_database(cleaned_records, conn)
        
        return jsonify({
            'status': 'success', 
            'msg': f'Đồng bộ thành công! Thêm mới: {inserted} đơn. Cập nhật: {updated} đơn.'
        })

    except Exception as e:
        # Bắt mọi lỗi (File sai định dạng, SQL lỗi...)
        return jsonify({'status': 'error', 'msg': f'Lỗi hệ thống: {str(e)}'})
    
    finally:
        # 5. ĐÓNG KẾT NỐI (QUAN TRỌNG NHẤT)
        # Khối finally LUÔN LUÔN CHẠY dù try thành công hay thất bại (có Exception)
        if conn:
            conn.close() 

@alloc_run_bp.route('/api/sap/get_so_list', methods=['GET'])
def get_sap_so_list():
    conn = None
    try:
        conn = db.get_connection()
        query = """
        SELECT DISTINCT 
            [Sales Document] as so_number, 
            [Customer] as customer_name,
            [PO.] as po_number
        FROM [factory].[dbo].[so] WITH (NOLOCK)
        ORDER BY [Sales Document] DESC
        """
        cursor = conn.cursor()
        cursor.execute(query)
        rows = db.fetchall_as_dict(cursor) # Hàm helper của bạn
        return jsonify({'status': 'success', 'data': rows})
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
# --- API MỚI: Lấy chi tiết Material và Parse thông tin ---
@alloc_run_bp.route('/api/sap/get_so_items', methods=['POST'])
def get_sap_so_items():
    conn = None
    try:
        so_number = request.json.get('so_number')
        if not so_number: return jsonify({'status':'error', 'msg':'Thiếu số SO'})

        conn = db.get_connection()
        # Lấy dữ liệu theo SO
        query = """
        SELECT 
            [Material] as material_code,
            [Item Description] as description,
            [Quantity (KG)] as total_qty,
            [Shipped Quantity (KG)] as shipped_qty,
            [Customer] as customer_name
        FROM [factory].[dbo].[so] WITH (NOLOCK)
        WHERE [Sales Document] = ?
        """
        cursor = conn.cursor()
        cursor.execute(query, (so_number,))
        rows = db.fetchall_as_dict(cursor)
        # --- LOGIC PARSE ITEM DESCRIPTION ---
        results = []
        for r in rows:
            desc = r['description']
            thick = 0.0
            width = 0.0
            grade = ""
            dim_match = re.search(r'(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)', desc)
            
            if dim_match:
                try:
                    thick = float(dim_match.group(1))
                    width = float(dim_match.group(2))
                    
                    # 2. Logic đoán Mác thép (Grade)
                    # Thường Mác thép nằm sau kích thước hoặc ở cuối chuỗi
                    # Cách đơn giản: Lấy phần chuỗi SAU kích thước và làm sạch
                    remaining = desc[dim_match.end():].strip()
                    # Tách các từ, thường mác thép là từ đầu tiên sau kích thước (VD: SAE1006)
                    if remaining:
                        parts = remaining.split()
                        grade = parts[0] # Lấy từ đầu tiên (SPHC, SAE1006, SS400...)
                except:
                    pass

            # Tính khối lượng cần (Còn lại)
            # needed_qty = float(r['total_qty'] or 0) - float(r['shipped_qty'] or 0)
            needed_qty = float(r['total_qty'] or 0)
            if needed_qty < 0: needed_qty = 0

            results.append({
                'material_code': r['material_code'],
                'description': desc,
                'thick': thick,
                'width': width,
                'grade': grade, # Mác thép parse được
                'customer_name': r['customer_name'],
                'req_weight': needed_qty # Tự động trừ đi lượng đã ship
            })

        return jsonify({'status': 'success', 'data': results})
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
@alloc_run_bp.route('/allocation_run', methods=['GET'])
@permission_required('allocation_run')
def allocation_run_page():
    """Trang Chạy Phân bổ"""
    return render_template('allocation_run.html')

@alloc_run_bp.route('/api/run_batch_allocation', methods=['POST'])
def run_batch_allocation():
    conn = None
    try:
        req_items = request.json.get('items', [])
        # [MỚI] Nhận tham số mode từ frontend (Mặc định là ALLOCATE)
        mode = request.json.get('mode', 'ALLOCATE')
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # 1. Lấy dữ liệu TDC Master & Version (GIỮ NGUYÊN)
        query_tdc = """
            SELECT m.id as master_id, m.tdc_code, m.customer_name, m.grade, 
                v.criteria_json, v.id as version_id
            FROM tdc_master m
            JOIN tdc_versions v ON m.id = v.master_id
            WHERE v.status = 'Active'
            AND CAST(GETDATE() AS DATE) >= v.valid_from 
            AND (v.valid_to IS NULL OR CAST(GETDATE() AS DATE) <= v.valid_to) 
        """
        cursor.execute(query_tdc)
        active_tdcs = db.fetchall_as_dict(cursor)
        tdc_map = {row['master_id']: dict(row) for row in active_tdcs}

        # 2. Pre-filter: Lấy danh sách SO và Factory (GIỮ NGUYÊN)
        so_list_for_query = set()
        for req in req_items:
            so_num = str(req.get('so_number', '')).strip()
            if so_num and so_num != 'N/A':
                so_list_for_query.add(so_num)
        
        so_factory_map = {}
        if so_list_for_query:
            placeholders = ','.join(['?'] * len(so_list_for_query))
            query_factory = f"""
                SELECT [Sales Document] as so, [Material] as mat, 
                CASE 
                    WHEN [Factory] = N'Hòa Phát Dung Quất' THEN 'HRC1'
                    WHEN [Factory] = N'Hòa Phát Dung Quất 2' THEN 'HRC2'
                    ELSE [Factory] 
                END as factory 
                FROM [factory].[dbo].[so] WITH (NOLOCK)
                WHERE [Sales Document] IN ({placeholders})
            """
            cursor.execute(query_factory, tuple(so_list_for_query))
            rows_fac = db.fetchall_as_dict(cursor)
            
            for r in rows_fac:
                k = (str(r['so']).strip(), str(r['mat']).strip())
                so_factory_map[k] = str(r['factory']).strip()
            
            print(f"🏭 Loaded Factory info for {len(so_factory_map)} SO items.")

        # ==============================================================================
        # 3. LẤY TỒN KHO (TÙY CHỈNH THEO MODE)
        # ==============================================================================
        
        # ==============================================================================
        # 3. LẤY TỒN KHO (TÙY CHỈNH THEO MODE)
        # ==============================================================================
        inventory_source = {} 

        if mode == 'CHECK':
            # --- [MODE CHECK GIỮ NGUYÊN NHƯ CŨ CỦA BẠN] ---
            if not so_list_for_query: return jsonify({'status': 'error', 'msg': 'Cần có số SO để kiểm tra mapping!'})
            mat_list_for_query = set([str(req.get('material_code', '')).strip() for req in req_items if req.get('material_code')])
            
            so_placeholders = ','.join(['?'] * len(so_list_for_query))
            params = list(so_list_for_query)
            query_mapped = f"""
                SELECT 
                    c.coil_id, c.ID_XuLy, c.scores, c.grade, c.weight, 
                    c.target_thick, c.target_width, c.factory, c.Nhom, c.production_date, c.TARGET_LV2,
                    c.prime_status, c.mapped_po, c.qc_msg,
                    CAST(k.[SO Mapping] AS VARCHAR) as mapped_so,
                    CAST(k.[Material] AS VARCHAR) as mapped_material
                FROM coil_data c WITH (NOLOCK)
                INNER JOIN kho k WITH (NOLOCK) ON k.[ID Cuộn Bó] IN (c.coil_id, c.ID_XuLy)
                WHERE k.[SO Mapping] IN ({so_placeholders}) AND c.production_date >= '2026-02-01'
            """
            if mat_list_for_query:
                query_mapped += f" AND CAST(k.[Material] AS NVARCHAR(50)) IN ({','.join(['?'] * len(mat_list_for_query))})"
                params.extend(list(mat_list_for_query))

            cursor.execute(query_mapped, tuple(params))
            rows = db.fetchall_as_dict(cursor)
            
            for r in rows:
                so_key = str(r['mapped_so']).strip()
                if so_key not in inventory_source: inventory_source[so_key] = []
                item = dict(r)
                item['scores'] = json.loads(item['scores']) if item['scores'] else {}
                item['mapped_po'] = str(r['mapped_po']).strip() if r.get('mapped_po') else '0'
                item['prime_status'] = str(r['prime_status']).strip() if r.get('prime_status') else 'PRIME'
                item['weight'] = float(item['weight'] or 0.0)
                item['thick'] = float(item['target_thick'] or 0.0)
                item['width'] = float(item['target_width'] or 0.0)
                inventory_source[so_key].append(item)
                
            print(f"🕵️ CHECK MODE: Found {len(rows)} mapped coils.")

        else:
            # --- [MODE ALLOCATE MỚI]: Gộp Hàng Tồn (0) và Hàng MTO (SO) ---
            req_grades = set()
            req_sos = set()
            for req in req_items:
                tdc_info = tdc_map.get(int(req.get('tdc_id', 0)))
                if tdc_info: req_grades.add(tdc_info['grade'])
                so_num = str(req.get('so_number', '')).strip()
                if so_num and so_num != 'N/A': req_sos.add(so_num)
            
            if not req_grades: return jsonify({'status': 'error', 'msg': 'Không có grade nào hợp lệ'})
            
            # Xây dựng điều kiện mapped_po
            params = list(req_grades)
            so_condition = "AND (c.mapped_po IS NULL OR c.mapped_po = '' OR c.mapped_po = '0')"
            if req_sos:
                so_placeholders = ','.join(['?'] * len(req_sos))
                so_condition = f"AND (c.mapped_po IS NULL OR c.mapped_po = '' OR c.mapped_po = '0' OR c.mapped_po IN ({so_placeholders}))"
                params.extend(list(req_sos))

            query_inventory = f"""
                SELECT 
                    c.coil_id, c.ID_XuLy, c.scores, c.grade, c.weight, 
                    c.target_thick, c.target_width, c.factory, c.Nhom, c.production_date, c.TARGET_LV2,
                    c.mapped_po, c.prime_status,c.Nhom,c.qc_msg
                FROM coil_data c WITH (NOLOCK)
                WHERE   c.production_date >= '2026-07-01 08:00:00'
                    AND c.sap_so_mapping IS NULL AND c.allocated_order IS NULL
                    AND c.grade IN ({','.join(['?'] * len(req_grades))})
                    {so_condition}
                    AND EXISTS (
                        SELECT 1 FROM [factory].[dbo].[sanluong] sl WITH (NOLOCK)
                        WHERE sl.[ID Cuộn Bó] IN (c.coil_id, c.ID_XuLy) AND sl.[Đã nhập kho] = 'Yes' AND (sl.[Đã xuất kho] IS NULL OR sl.[Đã xuất kho] = '')
                    )       
            """
            cursor.execute(query_inventory, tuple(params))
            rows = db.fetchall_as_dict(cursor)
            
            for r in rows:
                grade_key = str(r['grade']).strip().upper()
                if grade_key not in inventory_source: inventory_source[grade_key] = []
                item = dict(r)
                item['scores'] = json.loads(item['scores']) if item['scores'] else {}
                item['mapped_po'] = str(r['mapped_po']).strip() if r.get('mapped_po') else '0'
                item['prime_status'] = str(r['prime_status']).strip() if r.get('prime_status') else 'PRIME'
                item['weight'] = float(item['weight'] or 0.0)
                item['thick'] = float(item['target_thick'] or 0.0)
                item['width'] = float(item['target_width'] or 0.0)
                inventory_source[grade_key].append(item)
            
            print(f"📦 ALLOC MODE: Loaded {len(rows)} coils (MTO + MTS).")

        # ==============================================================================
        # 4. BỘ NÃO PHÂN LOẠI & TÍNH ĐIỂM (CORE LOGIC)
        # ==============================================================================
        import time
        start_time = time.time()
        all_potential_candidates = [] 
        req_progress = { i: {'filled': 0.0, 'target': float(req.get('req_weight', 0))} for i, req in enumerate(req_items) }

        for req_idx, req in enumerate(req_items):
            try:
                so_number = str(req.get('so_number', 'N/A')).strip()
                material_code = str(req.get('material_code', '')).strip()
                target_factory = so_factory_map.get((so_number, material_code))
                tdc_id = int(req.get('tdc_id', 0))
                
                req_thick = float(req.get('thick', 0))
                req_alloc_thick = float(req.get('alloc_thick') or req_thick)
                req_width = float(req.get('width', 0))
                req_min_w = float(req.get('min_weight', 0))
                req_max_w = float(req.get('max_weight', 0))

                tdc_info = tdc_map.get(tdc_id)
                if not tdc_info: continue
                criteria_list = json.loads(tdc_info['criteria_json']) if tdc_info['criteria_json'] else []
                TOTAL_CRITERIA_COUNT = len(criteria_list)
                
                candidates_to_check = inventory_source.get(so_number, []) if mode == 'CHECK' else inventory_source.get(tdc_info['grade'], [])
                if not candidates_to_check: continue

                for coil in candidates_to_check:
                    # --- 1. LỌC NHÀ MÁY ---
                    if target_factory and coil['factory'] != target_factory: continue

                    # --- 2. LỌC KÍCH THƯỚC (BẢO TOÀN LOGIC CÁN ÂM CỦA BẠN) ---
                    is_negative_tolerance = (req_alloc_thick != req_thick)
                    if is_negative_tolerance:
                        coil_actual_thick = float(coil.get('TARGET_LV2') or coil['thick'])
                        if abs(coil_actual_thick - req_alloc_thick) > 0.02: continue
                    else:
                        if abs(coil['thick'] - req_thick) > 0.00: continue 

                    if abs(coil['width'] - req_width) > 0.0: continue
                    if req_min_w > 0 and coil['weight'] < req_min_w: continue
                    if req_max_w > 0 and coil['weight'] > req_max_w: continue

                    # --- 3. ĐỊNH TUYẾN MTO/MTS & TÍNH ĐIỂM ---
                    total_penalty = 0
                    failed_reasons = [] 
                    match_type = 'PERFECT'
                    eval_class = ''
                    sort_priority_list = []
                    is_rejected = False
                    met_count = 0

                    coil_mpo = coil.get('mapped_po', '0')

                    if mode == 'ALLOCATE':
                        if coil_mpo != '0' and coil_mpo != so_number: 
                            continue # Hàng của SO khác -> Bỏ qua
                            
                        if coil_mpo == so_number:
                            # 🥇 NHÓM MTO CHÍNH CHỦ
                            met_count = TOTAL_CRITERIA_COUNT
                            if coil.get('prime_status', 'PRIME') == 'PRIME':
                                eval_class = 'MTO_PRIME'
                                sort_priority_list = (1, 0, coil.get('production_date', '9999'))
                            else:
                                eval_class = 'MTO_NON_PRIME'
                                sort_priority_list = (2, 0, coil.get('production_date', '9999'))
                                msg = str(coil.get('qc_msg') or '').strip()
                                if msg:
                                    failed_reasons.append(f"{msg}")
                        else:
                            # 🥈 NHÓM TỒN KHO MTS (Phải chấm điểm với TDC)
                            for idx, crit in enumerate(criteria_list):
                                defect_key = crit['defect']
                                val = coil['scores'].get(defect_key, 0)
                                allowed_range = crit.get('range', [])
                                
                                if val == 0: 
                                    total_penalty += ((TOTAL_CRITERIA_COUNT - idx) * 25)
                                    failed_reasons.append(f"{crit.get('name_vi', defect_key)}:Thiếu")
                                elif val in allowed_range:
                                    met_count += 1
                                else:
                                    dist = abs(val - (min(allowed_range, key=lambda x: abs(x - val)) if allowed_range else 1))
                                    total_penalty += ((TOTAL_CRITERIA_COUNT - idx) * dist * 5)
                                    failed_reasons.append(f"{crit.get('name_vi', defect_key)}:C{val}(Lệch {dist})")

                            if total_penalty == 0:
                                eval_class = 'MTS_PASS'
                                sort_priority_list = (3, coil.get('production_date', '9999'), 0)
                            else:
                                eval_class = 'MTS_FAIL'
                                match_type = 'PROP_MISMATCH'
                                sort_priority_list = (4, total_penalty, coil.get('production_date', '9999'))
                    else:
                        # CHẾ ĐỘ CHECK (Giữ nguyên cho form SAP cũ)
                        # ... (Ở đây bạn có thể nhúng lại logic check cũ nếu cần, hoặc mặc định như trên)
                        eval_class = 'CHECKED'
                        sort_priority_list = (99, 0, 0)

                    if is_rejected: continue

                    # --- 4. ĐÓNG GÓI KẾT QUẢ ---
                    c_item = coil.copy()
                    c_item.update({
                        'req_index': req_idx, 
                        'eval_class': eval_class,
                        'sort_priority': sort_priority_list,
                        'penalty': total_penalty,
                        'match_type': match_type,
                        'customer_alloc': tdc_info['customer_name'],
                        'so_number': so_number,
                        'material_code': material_code,
                        'match_ratio': f"{met_count}/{TOTAL_CRITERIA_COUNT}", 
                        'match_pct': round((met_count / TOTAL_CRITERIA_COUNT) * 100, 1) if TOTAL_CRITERIA_COUNT > 0 else 0,
                        'failed_msg': ', '.join(failed_reasons),
                        'tdc_difficulty': TOTAL_CRITERIA_COUNT ,
                        'tdc_id': tdc_id,
                        'tdc_version_id': tdc_info['version_id'],
                        'group_name': str(coil.get('Nhom', 'N/A')).strip() if coil.get('Nhom') else 'N/A'
                    })
                    all_potential_candidates.append(c_item)

            except Exception as ex: 
                print(f"Lỗi Alloc Item {req_idx}: {ex}")
                continue

        # ==============================================================================
        # 5. SẮP XẾP & TRẢ KẾT QUẢ
        # ==============================================================================
        if mode == 'CHECK':
            all_potential_candidates.sort(key=lambda x: (x['so_number'], x['material_code'], x['penalty']))
            return jsonify({'status': 'success', 'allocated': all_potential_candidates})

        # 🌟 MAGIC HAPPENS HERE: Thuật toán tham lam kết hợp Sort Đa tầng
        all_potential_candidates.sort(key=lambda x: (
            x['so_number'], 
            x['material_code'], 
            x['sort_priority'] # (1,0,date) -> (2,0,date) -> (3,date,0) -> (4,penalty,date)
        ))
        
        final_results = []
        used_coil_ids = set()

        for cand in all_potential_candidates:
            if cand['coil_id'] in used_coil_ids: continue
            progress = req_progress[cand['req_index']]
            if progress['filled'] >= progress['target']: continue

            final_results.append(cand)
            used_coil_ids.add(cand['coil_id'])
            progress['filled'] += cand['weight']
        
        elapsed_time = time.time() - start_time
        print(f"⏱️ Allocation: {len(final_results)} items in {elapsed_time:.3f}s")
        return jsonify({'status': 'success', 'allocated': final_results})

    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()

@alloc_run_bp.route('/api/confirm_multi_so', methods=['POST'])
def confirm_multi_so():
    conn = None
    try:
        req = request.json
        alloc_list = req.get('allocations', [])
        so_meta_list = req.get('so_meta', [])
        current_user = session.get('username', 'Unknown')
        if not alloc_list:
            return jsonify({'status': 'error', 'msg': 'Danh sách phân bổ trống!'})

        conn = db.get_connection()
        cursor = conn.cursor()
        
        import datetime
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # --- GIAI ĐOẠN 1: CẬP NHẬT CẤU TRÚC ĐƠN HÀNG (HEADER & DETAILS) ---
        for meta in so_meta_list:
            so_raw = str(meta.get('so_number')).strip()
            try:
                so_number = int(so_raw)
            except:
                so_number = so_raw

            cust_name = meta.get('customer_name')
            material_code = str(meta.get('material', '')).strip()
            
            # 1. XỬ LÝ HEADER (sales_orders)
            cursor.execute("SELECT 1 FROM sales_orders WHERE so_number = ?", (so_number,))
            if not cursor.fetchone():
                # Chưa có -> Tạo mới
                cursor.execute("""
                    INSERT INTO sales_orders (so_number, customer_name, order_date, created_at, status)
                    VALUES (?, ?, ?, ?, 'Processing')
                """, (so_number, cust_name, now, now))
            else:
                # Đã có -> Cập nhật trạng thái sang Processing
                cursor.execute("UPDATE sales_orders SET status = 'Processing' WHERE so_number = ?", (so_number,))

            # 2. XỬ LÝ DETAIL (so_details)
            # Kiểm tra dòng này đã có chưa (ưu tiên theo material code)
            check_sql = "SELECT id FROM so_details WHERE so_number = ? AND material_code = ?"
            cursor.execute(check_sql, (so_number, material_code))
            min_w = meta.get('min_weight', 0)
            max_w = meta.get('max_weight', 0)
            row_tdc_id = meta.get('tdc_id')
            if row_tdc_id == '': row_tdc_id = None

            if not cursor.fetchone():
                # [THÊM min_weight, max_weight VÀO LỆNH INSERT]
                cursor.execute("""
                    INSERT INTO so_details (
                        so_number, material_code, description, grade, 
                        thickness, alloc_thick, width, total_weight, tdc_id, status, min_weight, max_weight
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Processing', ?, ?)             
                """, (
                    so_number, material_code,
                    f"{meta.get('grade')} {meta.get('thick')}x{meta.get('width')}",
                    meta.get('grade', ''), float(meta.get('thick', 0)), float(meta.get('thick', 0)), 
                    float(meta.get('width', 0)), float(meta.get('req_weight', 0)),
                    row_tdc_id, min_w, max_w
                ))
            else:
                # [THÊM min_weight, max_weight VÀO LỆNH UPDATE]
                if row_tdc_id is not None:
                    cursor.execute("""
                        UPDATE so_details SET tdc_id = ?, min_weight = ?, max_weight = ? 
                        WHERE so_number = ? AND material_code = ?
                    """, (row_tdc_id, min_w, max_w, so_number, material_code))
                else:
                    cursor.execute("""
                        UPDATE so_details SET min_weight = ?, max_weight = ? 
                        WHERE so_number = ? AND material_code = ?
                    """, (min_w, max_w, so_number, material_code))

        # --- GIAI ĐOẠN 2: CẬP NHẬT KHO (coil_data) ---
        success_count = 0
        allocated_so_set = set()

        for item in alloc_list:
            so = str(item['so_number']).strip()
            cid = item['coil_id']
            user_note = item.get('note', '')
            p_penalty = item.get('penalty', 0)
            p_pct = item.get('match_pct', 0)
            p_ratio = item.get('match_ratio', '')
            p_msg = item.get('failed_msg', '')
            alloc_tdc_id = item.get('tdc_id')
            allocated_so_set.add(so)

            # Tìm thông tin để update vào cuộn
            found_meta = next((m for m in so_meta_list if str(m['so_number']) == so), {})
            cust_name_alloc = found_meta.get('customer_name', '')
            mat_code = str(item.get('material_code', '')).strip()
            if not mat_code:
                mat_code = found_meta.get('material', '')

            # Fallback tên khách
            if not cust_name_alloc:
                cursor.execute("SELECT customer_name FROM sales_orders WHERE so_number = ?", (so,))
                res = cursor.fetchone()
                cust_name_alloc = res[0] if res else "Unknown"

            # 1. Update Coil Data
            update_sql = """
                UPDATE coil_data 
                SET allocated_to = ?, allocated_order = ?, allocated_material = ?, 
                    allocated_at = ?, alloc_note = ?,
                    alloc_penalty = ?, alloc_match_pct = ?, alloc_match_ratio = ?, 
                    alloc_failed_msg = ?, allocated_tdc_id = ? 
                WHERE coil_id = ? AND (allocated_to IS NULL OR allocated_to = '')
            """
            cursor.execute(update_sql, (
                cust_name_alloc, so, mat_code, now, user_note, 
                p_penalty, p_pct, p_ratio, p_msg, alloc_tdc_id,
                cid
            ))
            
            if cursor.rowcount > 0:
                success_count += 1
                
                # 2. [MỚI] GHI LOG CHI TIẾT NGAY TẠI ĐÂY
                # Lưu chi tiết: Ai đã gán cuộn nào vào SO nào
                log_desc = f"Gán cuộn {cid} ({cust_name_alloc}) cho đơn {so}. Ghi chú: {user_note}"
                
                cursor.execute("""
                    INSERT INTO action_history (action_type, ref_id, description, user_name, created_at)
                    VALUES (?, ?, ?, ?, ?)
                """, ('ALLOCATE_ITEM', so, log_desc, current_user, now))
        
        conn.commit()
        return jsonify({
            'status': 'success', 
            'msg': f'Thành công! Đã gán {success_count} cuộn vào {len(allocated_so_set)} đơn hàng.'
        })

    except Exception as e:
        if conn: conn.rollback()
        print(f"ROLLBACK: {str(e)}")
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
# 1. API Lấy danh sách Backlog (Đơn hàng đang chờ xử lý)
@alloc_run_bp.route('/api/get_backlog', methods=['GET'])
def get_backlog():
    conn = None
    try:
        conn = db.get_connection()
        
        # THÊM ĐIỀU KIỆN: WHERE d.status = 'Pending' HOẶC d.status IS NULL
        query = """
        SELECT 
            d.*, 
            s.customer_name,
            m.tdc_code,
            v.version_no,
            v.status as v_status,
            (SELECT ISNULL(SUM(weight), 0) FROM coil_data c 
            WHERE c.allocated_order = d.so_number AND c.allocated_material = d.material_code) as allocated_actual
        FROM so_details d
        LEFT JOIN sales_orders s ON d.so_number = s.so_number
        LEFT JOIN tdc_master m ON d.tdc_id = m.id
        LEFT JOIN tdc_versions v ON m.id = v.master_id AND v.status = 'Active'
        
        -- [QUAN TRỌNG] Chỉ lấy dòng đang Pending
        WHERE d.status = 'Pending'
            AND s.status IN ('Pending', 'Processing')
        
        ORDER BY d.id DESC
        """
        cursor = conn.cursor()
        cursor.execute(query)
        rows = db.fetchall_as_dict(cursor)
        
        return jsonify({'status': 'success', 'data': rows})
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()

# 2. API Lưu/Cập nhật Item vào Backlog (Auto-save)
@alloc_run_bp.route('/api/save_backlog_item', methods=['POST'])
def save_backlog_item():
    conn = None
    try:
        req = request.json
        so_raw = req.get('so_number')
        
        # --- [FIX BỌC THÉP BIGINT] ---
        try:
            so = int(str(so_raw).strip())
        except (ValueError, TypeError):
            return jsonify({'status': 'error', 'msg': f'Số SO "{so_raw}" không hợp lệ. Hệ thống BIGINT yêu cầu SO bắt buộc phải là SỐ.'})

        mat = str(req.get('material_code', '')).strip()
        desc = req.get('description', '')
        cust = req.get('customer_name')
        grade = req.get('grade')
        
        # Dùng `or 0` để tránh lỗi sập khi UI gửi lên null hoặc rỗng
        thick = float(req.get('thick') or 0)
        width = float(req.get('width') or 0)
        alloc_thick_raw = req.get('alloc_thick')
        alloc_thick = float(alloc_thick_raw) if alloc_thick_raw else thick
        weight = float(req.get('req_weight') or 0)
        min_w = float(req.get('min_weight') or 0)
        max_w = float(req.get('max_weight') or 0)
        
        tdc_id = req.get('tdc_id')
        if not tdc_id or str(tdc_id).strip() == "": 
            tdc_id = None

        conn = db.get_connection()
        cursor = conn.cursor()

        # B1: Đảm bảo Header (sales_orders) tồn tại (ĐÃ FIX SELECT ĐÚNG BẢNG)
        cursor.execute("SELECT 1 FROM sales_orders WHERE so_number = ?", (so,))
        if not cursor.fetchone():
            import datetime
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cursor.execute("INSERT INTO sales_orders (so_number, customer_name, created_at, status) VALUES (?, ?, ?, 'Pending')", (so, cust, now))

        # B2: Kiểm tra xem dòng Detail này đã có chưa (ĐÃ XÓA CODE LẶP)
        cursor.execute("SELECT id FROM so_details WHERE so_number = ? AND material_code = ?", (so, mat))
        row = cursor.fetchone()

        item_id = None
        
        if row:
            item_id = row[0]
            cursor.execute("SELECT min_weight, max_weight, tdc_id, alloc_thick FROM so_details WHERE id = ?", (item_id,))
            old_data = cursor.fetchone()
            
            # [SỬA Ở ĐÂY]: Chỉ cần Frontend CÓ GỬI key này lên (dù là rỗng hay 0), ta ưu tiên lấy giá trị mới. 
            # Nếu Frontend KHÔNG GỬI (như lúc addFromSap), ta mới giữ lại old_data.
            
            final_min_w = float(req['min_weight'] or 0) if 'min_weight' in req else (old_data[0] or 0)
            final_max_w = float(req['max_weight'] or 0) if 'max_weight' in req else (old_data[1] or 0)
            
            if 'tdc_id' in req:
                req_tdc = req['tdc_id']
                final_tdc_id = req_tdc if str(req_tdc).strip() else None
            else:
                final_tdc_id = old_data[2]
                
            final_alloc_thick = float(req['alloc_thick'] or thick) if 'alloc_thick' in req else (old_data[3] or thick)

            cursor.execute("""
                UPDATE so_details 
                SET total_weight = ?, min_weight = ?, max_weight = ?, tdc_id = ?, 
                    thickness = ?, alloc_thick = ?, width = ?, grade = ?, status = 'Pending'
                WHERE id = ?
            """, (weight, final_min_w, final_max_w, final_tdc_id, thick, final_alloc_thick, width, grade, item_id))
            conn.commit()
            # Trả về dữ liệu đầy đủ để Frontend hiển thị
            return jsonify({
                'status': 'success', 
                'id': item_id, 
                'data': {
                    'min_w': final_min_w, 
                    'max_w': final_max_w, 
                    'tdc_id': final_tdc_id,
                    'alloc_thick': final_alloc_thick
                }
            })
        else:
            cursor.execute("""
                INSERT INTO so_details (so_number, material_code, description, grade, thickness, alloc_thick, width, total_weight, min_weight, max_weight, tdc_id, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pending')
            """, (so, mat, desc, grade, thick, alloc_thick, width, weight, min_w, max_w, tdc_id))
            
            try: 
                item_id = cursor.lastrowid 
            except: 
                pass

            if not item_id:
                cursor.execute("SELECT TOP 1 id FROM so_details WHERE so_number=? AND material_code=? ORDER BY id DESC", (so, mat))
                new_row = cursor.fetchone()
                if new_row: item_id = new_row[0]

        conn.commit()
        return jsonify({'status': 'success', 'id': item_id, 'msg': 'Đã lưu'})
        
    except Exception as e:
        if conn: conn.rollback() # Thêm rollback để an toàn khi lỗi
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()

# 3. API Xóa khỏi Backlog (Nút X)
@alloc_run_bp.route('/api/remove_backlog_item', methods=['POST'])
def remove_backlog_item():
    conn = None
    try:
        req = request.json
        item_id = req.get('id')
        
        conn = db.get_connection()
        # THAY VÌ DELETE, TA UPDATE TRẠNG THÁI
        if item_id:
            conn.execute("UPDATE so_details SET status = 'Hidden' WHERE id = ?", (item_id,))
        else:
            # Fallback
            conn.execute("UPDATE so_details SET status = 'Hidden' WHERE so_number = ? AND material_code = ?", 
                         (req.get('so'), req.get('material')))
        
        conn.commit()
        
        return jsonify({'status': 'success', 'msg': 'Đã xóa khỏi danh sách (Dữ liệu gốc vẫn được lưu)'})
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
@alloc_run_bp.route('/api/check_item_status', methods=['POST'])
def check_item_status():
    conn = None
    try:
        req = request.json
        so = req.get('so_number')
        mat = req.get('material_code')
        
        if not so or not mat: 
            return jsonify({'status': 'error', 'msg': 'Thiếu thông tin'})

        conn = db.get_connection()
        cursor = conn.cursor()

        # 1. Kiểm tra trong Backlog (Bảng so_details)
        # Xem người dùng đã từng thêm dòng này vào danh sách chờ chưa
        cursor.execute("""
            SELECT SUM(total_weight) 
            FROM so_details 
            WHERE so_number = ? AND material_code = ?
        """, (so, mat))
        row_backlog = cursor.fetchone()
        backlog_weight = row_backlog[0] if row_backlog and row_backlog[0] else 0

        # 2. Kiểm tra đã phân bổ xong (Bảng coil_data)
        # Xem thực tế đã gán được bao nhiêu tấn cho dòng này
        cursor.execute("""
            SELECT SUM(weight) 
            FROM coil_data 
            WHERE allocated_order = ? AND allocated_material = ?
        """, (so, mat))
        row_alloc = cursor.fetchone()
        allocated_weight = row_alloc[0] if row_alloc and row_alloc[0] else 0

    
        
        return jsonify({
            'status': 'success',
            'backlog_weight': backlog_weight,
            'allocated_weight': allocated_weight
        })
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
# RÀ SOÁT LẠI CUỘN CHO P.KH (CẬP NHẬT ĐỌC TỪ RAW_DATA)
@alloc_run_bp.route('/api/check_custom_coils', methods=['POST'])
def check_custom_coils():
    conn = None
    try:
        req = request.json
        raw_ids = req.get('coil_ids', [])
        grades = req.get('grades', [])  
        t_min_in = float(req.get('t_min') or 0)
        t_max_in = float(req.get('t_max') or 0)
        w_min_in = float(req.get('w_min') or 0)
        w_max_in = float(req.get('w_max') or 0)
        custom_props = req.get('custom_props', {})
        
        # Nhận biến TDC
        tdc_id = req.get('tdc_id')

        search_ids = [str(x).strip() for x in raw_ids if str(x).strip()]
        if not search_ids:
            return jsonify({'status': 'error', 'msg': 'Danh sách ID trống!'})

        conn = db.get_connection()
        cursor = conn.cursor()

        # 1. Trích xuất Tiêu chí TDC (nếu có)
        criteria_list = []
        if tdc_id and str(tdc_id).strip():
            query_tdc = """
                SELECT v.criteria_json 
                FROM tdc_versions v 
                WHERE v.master_id = ? AND v.status = 'Active'
            """
            cursor.execute(query_tdc, (tdc_id,))
            row_tdc = cursor.fetchone()
            if row_tdc and row_tdc[0]:
                criteria_list = json.loads(row_tdc[0])

        # 2. Truy vấn dữ liệu cuộn
        placeholders = ','.join(['?'] * len(search_ids))
        query = f"""
            SELECT coil_id, ID_XuLy, grade, target_thick, target_width, scores, raw_data, factory, Nhom
            FROM coil_data WITH (NOLOCK)
            WHERE coil_id IN ({placeholders}) OR ID_XuLy IN ({placeholders})
        """
        cursor.execute(query, search_ids + search_ids)
        rows = db.fetchall_as_dict(cursor)

        results = []
        found_input_ids = set() 

        for row in rows:
            coil_id = str(row['coil_id']).strip()
            skin_id = str(row['ID_XuLy']).strip() if row.get('ID_XuLy') else ''
            for s_id in search_ids:
                if s_id == coil_id or s_id == skin_id: found_input_ids.add(s_id)

            c_grade = str(row['grade']).strip().upper()
            c_thick = float(row['target_thick'] or 0)
            c_width = float(row['target_width'] or 0)
            
            try: c_scores = json.loads(row['scores']) if row['scores'] else {}
            except: c_scores = {}
            try: c_raw = json.loads(row['raw_data']) if row.get('raw_data') else {}
            except: c_raw = {}

            failed_reasons = []

            # --- KIỂM TRA ĐỘC LẬP: CÁC FILTER TRÊN GIAO DIỆN ---
            if grades and c_grade not in [g.upper() for g in grades]:
                failed_reasons.append(f"Mác thép ({c_grade} không thuộc {','.join(grades)})")
            
            # Kích thước
            if t_min_in > 0 or t_max_in > 0:
                c_t_min = float(c_raw.get('ThickMin', 0))
                c_t_max = float(c_raw.get('ThickMax', 0))
                if c_t_min == 0 and c_t_max == 0: failed_reasons.append("Thiếu data ThickMin/Max")
                else:
                    if t_min_in > 0 and c_t_min < t_min_in: failed_reasons.append(f"Dày Min < {t_min_in}")
                    if t_max_in > 0 and c_t_max > t_max_in: failed_reasons.append(f"Dày Max > {t_max_in}")

            if w_min_in > 0 or w_max_in > 0:
                c_w_min = float(c_raw.get('WidthMin', 0))
                c_w_max = float(c_raw.get('WidthMax', 0))
                if c_w_min == 0 and c_w_max == 0: failed_reasons.append("Thiếu data WidthMin/Max")
                else:
                    if w_min_in > 0 and c_w_min < w_min_in: failed_reasons.append(f"Rộng Min < {w_min_in}")
                    if w_max_in > 0 and c_w_max > w_max_in: failed_reasons.append(f"Rộng Max > {w_max_in}")

            # Hóa tính/Cơ tính gõ tay (TPHH thủ công)
            for prop_key, limits in custom_props.items():
                val = c_raw.get(prop_key)
                if val is None:
                    failed_reasons.append(f"Thiếu [{prop_key}]")
                    continue
                try: val = float(val)
                except: 
                    failed_reasons.append(f"[{prop_key}] không là số")
                    continue
                p_min = float(limits.get('min') or 0)
                p_max = float(limits.get('max') or 0)
                if p_min > 0 and val < p_min: failed_reasons.append(f"[{prop_key}] < {p_min}")
                if p_max > 0 and val > p_max: failed_reasons.append(f"[{prop_key}] > {p_max}")

            # --- KIỂM TRA ĐỘC LẬP: THEO TDC ĐƯỢC CHỌN ---
            if criteria_list:
                for crit in criteria_list:
                    defect_key = crit['defect']
                    val = c_scores.get(defect_key, 0)
                    allowed_range = crit.get('range', [])
                    
                    if val == 0: 
                        failed_reasons.append(f"TDC: Thiếu {crit.get('name_vi', defect_key)}")
                    elif val not in allowed_range:
                        dist = abs(val - (min(allowed_range, key=lambda x: abs(x - val)) if allowed_range else 1))
                        failed_reasons.append(f"TDC: {crit.get('name_vi', defect_key)} lệch {dist}")

            # Quyết định trạng thái
            status = 'FAIL' if failed_reasons else 'PASS'

            results.append({
                'search_id': coil_id, 'coil_id': coil_id, 'skin_id': skin_id,
                'grade': c_grade, 'thick': c_thick, 'width': c_width,
                'group': str(row['Nhom'] or 'N/A'),
                'status': status, 'reasons': failed_reasons
            })

        # Xử lý các ID không tìm thấy
        missing_ids = set(search_ids) - found_input_ids
        for m_id in missing_ids:
            results.append({
                'search_id': m_id, 'coil_id': m_id, 'skin_id': '', 'grade': '-', 
                'thick': '-', 'width': '-', 'group': '-',
                'status': 'NOT_FOUND', 'reasons': ['Không tìm thấy ID trong DB']
            })

        sort_order = {'FAIL': 1, 'NOT_FOUND': 2, 'PASS': 3}
        results.sort(key=lambda x: (sort_order.get(x['status'], 99), x['search_id']))

        return jsonify({'status': 'success', 'data': results})

    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
def clean_note_qc(raw_note):
    if not raw_note:
        return ""
    try:
        # Thử parse JSON
        note_dict = json.loads(raw_note)
        if isinstance(note_dict, dict):
            note_parts = []
            for key in ['surf', 'geo', 'prop', 'app']:
                val = str(note_dict.get(key, '')).strip()
                if val and val.lower() not in ['none', 'nan', 'null']:
                    note_parts.append(val)
            # Nối bằng dấu | nếu có data, ngược lại trả về rỗng
            return " | ".join(note_parts)
        return str(raw_note).strip()
    except Exception:
        # Nếu không phải JSON (text cũ nhập tay), giữ nguyên text
        return str(raw_note).strip()
@alloc_run_bp.route('/api/export_mtc', methods=['POST'])
def export_mtc():
    try:
        req = request.json
        raw_ids = req.get('coil_ids', [])
        search_ids = [str(x).strip() for x in raw_ids if str(x).strip()]
        if not search_ids:
            return jsonify({'status': 'error', 'msg': 'Danh sách ID trống!'}), 400

        conn = db.get_connection()
        cursor = conn.cursor()

        placeholders = ','.join(['?'] * len(search_ids))
        query = f"""
            SELECT coil_id, ID_XuLy, grade, target_thick, target_width, weight, 
                production_date, factory, raw_data, note_qc, TARGET_LV2, Speed,
                slab_grade_name, slab_grade, target_temp_coil
            FROM coil_data WITH (NOLOCK)
            WHERE coil_id IN ({placeholders}) OR ID_XuLy IN ({placeholders})
            ORDER BY production_date DESC
        """
        cursor.execute(query, search_ids + search_ids)
        rows = db.fetchall_as_dict(cursor)
        conn.close()

        if not rows:
            return jsonify({'status': 'error', 'msg': 'Không tìm thấy dữ liệu.'}), 404

        # Phân tách mảng cột theo yêu cầu cấu trúc mới
        INFO_COLS = ["STT", "ID Gốc", "Skin ID", "Mác Thép", "Mác phôi", "Mẻ phôi", "Kích thước", "Khối lượng (kg)", "Ngày SX"]
        MECH_MAP = {'YieldPoint': 'GH Chảy', 'Tensile': 'GH Bền', 'Elongation': 'Độ giãn', 'Hardness': 'Độ cứng'}
        CHEM_KEYS = ['C', 'Si', 'Mn', 'S', 'P', 'Cu', 'Ni', 'Cr', 'Mo', 'V', 'Ti', 'Al', 'Ca', 'B', 'Nb', 'CEV', 'O', 'N', 'H']
        
        GEOM_COLS = ["Dày TB", "Dày Min", "Dày Max", "Rộng TB", "Rộng Min", "Rộng Max"]
        TEMP_COLS = ["Nhiệt độ cài đặt", "Nhiệt độ Cuộn", "Nhiệt Min", "Nhiệt Max"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MTC_Export"

        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        row1, row2 = [], []

        # Tái cấu trúc chuỗi cột Header tầng 1 và tầng 2
        for h in INFO_COLS:
            row1.append("THÔNG TIN CHUNG")
            row2.append(h)
        for k, v in MECH_MAP.items():
            row1.append("CƠ TÍNH")
            row2.append(v)
        for k in CHEM_KEYS:
            row1.append("THÀNH PHẦN HÓA HỌC")
            row2.append(k)
        for h in GEOM_COLS:
            row1.append("KÍCH THƯỚC ĐO LƯỜNG")
            row2.append(h)
            
        row1.append("GHI CHÚ")
        row2.append("Ghi Chú QC")
        
        for h in TEMP_COLS:
            row1.append("NHIỆT ĐỘ CÁN")
            row2.append(h)

        ws.append(row1)
        ws.append(row2)

        def style_header(start_col, end_col, title, bg_color, row2_bg):
            if start_col < end_col:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            c1 = ws.cell(row=1, column=start_col)
            c1.value = title
            c1.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            c1.font = header_font
            c1.alignment = center_align
            c1.border = thin_border
            for c in range(start_col, end_col + 1):
                ws.cell(row=1, column=c).border = thin_border
                c2 = ws.cell(row=2, column=c)
                c2.fill = PatternFill(start_color=row2_bg, end_color=row2_bg, fill_type="solid")
                c2.font = Font(bold=True)
                c2.alignment = center_align
                c2.border = thin_border

        # Tính toán lại vị trí phân chia khối màu động
        idx_info_end = len(INFO_COLS)
        idx_mech_start = idx_info_end + 1
        idx_mech_end = idx_info_end + len(MECH_MAP)
        idx_chem_start = idx_mech_end + 1
        idx_chem_end = idx_mech_end + len(CHEM_KEYS)
        idx_geom_start = idx_chem_end + 1
        idx_geom_end = idx_chem_end + len(GEOM_COLS)
        idx_note = idx_geom_end + 1
        idx_temp_start = idx_note + 1
        idx_temp_end = idx_note + len(TEMP_COLS)

        style_header(1, idx_info_end, "THÔNG TIN CHUNG", "3B82F6", "DBEAFE")
        style_header(idx_mech_start, idx_mech_end, "CƠ TÍNH", "10B981", "D1FAE5")
        style_header(idx_chem_start, idx_chem_end, "THÀNH PHẦN HÓA HỌC", "F59E0B", "FEF3C7")
        style_header(idx_geom_start, idx_geom_end, "KÍCH THƯỚC ĐO LƯỜNG", "8B5CF6", "EDE9FE")
        style_header(idx_note, idx_note, "GHI CHÚ", "64748B", "F1F5F9")
        style_header(idx_temp_start, idx_temp_end, "NHIỆT ĐỘ CÁN", "EF4444", "FEE2E2")

        for idx, r in enumerate(rows):
            try: raw_data = json.loads(r.get('raw_data') or '{}')
            except: raw_data = {}

            # ==========================================
            # 1. LOGIC CHO SIZE (Kích thước danh định)
            # ==========================================
            target_thick = r.get('target_thick')
            if target_thick in [None, '', 0, 0.0, '0', '0.0']:
                target_thick = r.get('TARGET_LV2', '') 
            target_width = r.get('target_width', '')
            
            size = f"{target_thick} x {target_width}" if target_thick and target_width else ""
            prod_date = str(r.get('production_date', '')).split('.')[0]

            # ==========================================
            # 2. LOGIC CHO MEASUREMENTS (Đo lường thực tế)
            # ==========================================
            factory_id = r.get('factory', '')
            if factory_id == 'HRC1':
                thick_avg = r.get('TARGET_LV2', '')
            else:
                thick_avg = raw_data.get('Thick', '')
            
            # Chiều rộng thực tế bốc từ raw_data
            width_avg = raw_data.get('Width', '')

            # ------------------------------------------
            # ĐỔ DỮ LIỆU VÀO MẢNG ROW_DATA
            # ------------------------------------------
            # 1. Khối Thông tin chung (Dùng biến 'size' danh định)
            row_data = [
                idx + 1, 
                r.get('coil_id', ''), 
                r.get('ID_XuLy', ''), 
                r.get('grade', ''), 
                r.get('slab_grade_name', ''), 
                r.get('slab_grade', ''), 
                size, 
                r.get('weight', ''), 
                prod_date
            ]
            
            # 2. Khối Cơ tính
            for k in MECH_MAP.keys(): row_data.append(raw_data.get(k, ""))
            
            # 3. Khối Thành phần Hóa học
            for k in CHEM_KEYS: row_data.append(raw_data.get(k, ""))
            
            # 4. Khối Kích thước Đo lường (Dùng thick_avg và width_avg thực tế)
            row_data.extend([
                thick_avg, 
                raw_data.get('ThickMin', ''), 
                raw_data.get('ThickMax', ''), 
                width_avg, 
                raw_data.get('WidthMin', ''), 
                raw_data.get('WidthMax', '')
            ])
            
            # 5. Khối Ghi chú
            row_data.append(clean_note_qc(r.get('note_qc', '')))
            
            # 6. Khối Nhiệt độ
            row_data.extend([
                r.get('target_temp_coil', ''),        # Nhiệt độ cài đặt 
                r.get('Speed', ''),                   # Nhiệt độ cuộn thực tế
                raw_data.get('DcTempMin', ''), 
                raw_data.get('DcTempMax', '')
            ])

            ws.append(row_data)

            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=idx + 3, column=col_idx)
                cell.border = thin_border
                if col_idx == idx_note:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                else:
                    cell.alignment = center_align

        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx_note)].width = 35

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Export_MTC_HRC_Premium.xlsx')
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)}), 500

# --- ROUTE PREVIEW MTC CHỈNH SỬA LOGIC DỮ LIỆU ĐỘ DÀY THEO THIẾT KẾ NHÀ MÁY ---
@alloc_run_bp.route('/api/preview_mtc', methods=['POST'])
def preview_mtc():
    try:
        req = request.json
        raw_ids = req.get('coil_ids', [])
        search_ids = [str(x).strip() for x in raw_ids if str(x).strip()]
        if not search_ids:
            return jsonify({'status': 'error', 'msg': 'Danh sách ID trống!'}), 400

        conn = db.get_connection()
        cursor = conn.cursor()
        placeholders = ','.join(['?'] * len(search_ids))
        query = f"""
            SELECT coil_id, ID_XuLy, grade, target_thick, target_width, weight, 
                production_date, factory, raw_data, note_qc, TARGET_LV2, Speed,
                slab_grade_name, slab_grade, target_temp_coil
            FROM coil_data WITH (NOLOCK)
            WHERE coil_id IN ({placeholders}) OR ID_XuLy IN ({placeholders})
            ORDER BY production_date DESC
        """
        cursor.execute(query, search_ids + search_ids)
        rows = db.fetchall_as_dict(cursor)
        conn.close()

        if not rows:
            return jsonify({'status': 'error', 'msg': 'Không tìm thấy dữ liệu.'}), 404

        results = []
        MECH_MAP = ['YieldPoint', 'Tensile', 'Elongation', 'Hardness']
        CHEM_KEYS = ['C', 'Si', 'Mn', 'S', 'P', 'Cu', 'Ni', 'Cr', 'Mo', 'V', 'Ti', 'Al', 'Ca', 'B', 'Nb', 'CEV', 'O', 'N', 'H']

        for r in rows:
            try: raw_data = json.loads(r.get('raw_data') or '{}')
            except: raw_data = {}

            # ==========================================
            # 1. LOGIC CHO SIZE (Kích thước danh định)
            # ==========================================
            target_thick = r.get('target_thick')
            if target_thick in [None, '', 0, 0.0, '0', '0.0']:
                target_thick = r.get('TARGET_LV2', '') 
            target_width = r.get('target_width', '')
            
            size = f"{target_thick} x {target_width}" if target_thick and target_width else ""
            
            # ==========================================
            # 2. LOGIC CHO MEASUREMENTS (Đo lường thực tế)
            # ==========================================
            factory_id = r.get('factory', '')
            if factory_id == 'HRC1':
                thick_avg = r.get('TARGET_LV2', '')
            else:
                thick_avg = raw_data.get('Thick', '')
                
            width_avg = raw_data.get('Width', '')

            item = {
                'coil_id': r.get('coil_id', ''),
                'skin_id': r.get('ID_XuLy', ''),
                'grade': r.get('grade', ''),
                'slab_grade_name': r.get('slab_grade_name', ''), 
                'slab_grade': r.get('slab_grade', ''),           
                'size': size, # <= Truyền Kích thước danh định
                'weight': r.get('weight', ''),
                'note_qc': clean_note_qc(r.get('note_qc', '')),
                
                # <= Truyền Kích thước thực tế vào Measurements
                'measurements': {
                    'thick_avg': thick_avg,
                    'thick_min': raw_data.get('ThickMin', ''),
                    'thick_max': raw_data.get('ThickMax', ''),
                    'width_avg': width_avg,
                    'width_min': raw_data.get('WidthMin', ''),
                    'width_max': raw_data.get('WidthMax', ''),
                    'target_temp_coil': r.get('target_temp_coil', ''), 
                    'temp_coil': r.get('Speed', ''),
                    'temp_min': raw_data.get('DcTempMin', ''),
                    'temp_max': raw_data.get('DcTempMax', '')
                },
                'mech': {k: raw_data.get(k, '') for k in MECH_MAP},
                'chem': {k: raw_data.get(k, '') for k in CHEM_KEYS}
            }
            results.append(item)
        return jsonify({'status': 'success', 'data': results})
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)}), 500