from flask import Blueprint, app, render_template, current_app, request, jsonify, send_file, session,redirect, url_for
import pandas as pd
import threading
import json
import db  # Module db.py
import io
import requests
from apscheduler.schedulers.background import BackgroundScheduler
from utils.common import sanitize_data, standardize_id
from utils.scoring import process_coil_scores,get_all_grade_configs,calculate_metric_surface,calculate_metric_value, build_matrix_tooltip
from utils.sync_worker import sync_surface_defects, sync_properties_mysql,process_and_save_geometry,API_GEO_SINGLE_URL,process_coil_scores,LAST_DATA_UPDATE, FACTORY_CONFIGS
import time
import datetime
import re
dashboard_bp = Blueprint('dashboard_bp', __name__)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from auth.decorator import login_required, permission_required
dashboard_bp = Blueprint('dashboard_bp', __name__)
@dashboard_bp.route('/api/export_excel_qc', methods=['POST'])
def export_excel_qc():
    try:
        req = request.json
        factory = req.get('factory', 'ALL')
        start_date = req.get('start_date', '')
        end_date = req.get('end_date', '')
        coil_ids = req.get('coil_ids', [])
        if coil_ids and len(coil_ids) > 2000:
            return jsonify({'status': 'error', 'msg': 'Vui lòng chọn xuất tối đa 2000 cuộn mỗi lần khi tìm theo ID.'}), 400

        if not coil_ids and start_date and end_date:
            try:
                fmt_start = '%Y-%m-%dT%H:%M' if 'T' in start_date else '%Y-%m-%d'
                fmt_end = '%Y-%m-%dT%H:%M' if 'T' in end_date else '%Y-%m-%d'
                d_start = datetime.datetime.strptime(start_date, fmt_start)
                d_end = datetime.datetime.strptime(end_date, fmt_end)
                if (d_end - d_start).days > 31:
                    return jsonify({'status': 'error', 'msg': 'Dữ liệu quá lớn. Vui lòng chọn khoảng thời gian tối đa là 1 tháng.'}), 400
            except Exception:
                pass
        conn = db.get_connection()
        cursor = conn.cursor()

        base_select = """
            SELECT c.coil_id, c.ID_XuLy, c.grade, c.scores, c.weight, c.target_thick, c.target_width, 
                   c.production_date, c.factory, c.quality_class, c.note_qc,
                   c.[Order], c.TieuChuan, c.Ca,
                   c.prime_status, c.qc_status, c.mapped_po, c.rework_status, c.qc_msg,
                   tm.tdc_code,
                   sl.[Ngày sản xuất] AS sl_production_date
            FROM coil_data c WITH (NOLOCK)
            LEFT JOIN tdc_versions tv WITH (NOLOCK) ON c.target_tdc_version_id = tv.id
            LEFT JOIN tdc_master tm WITH (NOLOCK) ON tv.master_id = tm.id
            LEFT JOIN sanluong sl WITH (NOLOCK) ON c.coil_id = sl.[ID Cuộn Bó]
            WHERE 1=1
        """
        
        params = []

        if coil_ids:
            # --- TRƯỜNG HỢP TÌM THEO ID: BỎ QUA NHÀ MÁY & NGÀY THÁNG ---
            placeholders = ', '.join(['?'] * len(coil_ids))
            
            # Thêm alias c. vào điều kiện IN
            sql_part1 = f"{base_select} AND c.coil_id IN ({placeholders})"
            sql_part2 = f"{base_select} AND c.ID_XuLy IN ({placeholders})"
            
            # Gộp lại và sắp xếp
            sql = f"{sql_part1} UNION {sql_part2} ORDER BY production_date DESC"
            
            params.extend(coil_ids)
            params.extend(coil_ids)
            
        else:
            # --- TRƯỜNG HỢP TÌM THEO NGÀY THÁNG VÀ NHÀ MÁY ---
            sql = base_select
            
            if factory != 'ALL':
                sql += " AND c.factory = ?"
                params.append(factory)
                
            if start_date:
                s_str = start_date.replace('T', ' ')
                if len(s_str) == 10: s_str += " 00:00:00"
                elif len(s_str) == 16: s_str += ":00"
                sql += " AND c.production_date >= ?"
                params.append(s_str)
            if end_date:
                e_str = end_date.replace('T', ' ')
                if len(e_str) == 10: e_str += " 23:59:59.999"
                elif len(e_str) == 16: e_str += ":59.999"
                sql += " AND c.production_date <= ?"
                params.append(e_str)
                
            sql += " ORDER BY c.production_date DESC"

        cursor.execute(sql, tuple(params))
        rows = db.fetchall_as_dict(cursor)
        conn.close()

        if not rows:
            return jsonify({'status': 'error', 'msg': 'Không có dữ liệu trong khoảng thời gian hoặc ID đã chọn.'}), 404

        # 2. CẤU HÌNH TỪ ĐIỂN VÀ NHÓM LỖI
        DEFECT_MAP = {
            'SURFACE': {
                'oil': 'Gấp nếp', 'rust': 'Nếp nhăn', 'scratch_m': 'Vết hằn Pinch Roll', 'dirt': 'Gãy mặt',
                'other_s': 'Xỉ muối tiêu', 'gianbien': 'Giãn biên/Bụng', 'chambi': 'Chấm bi',
                'MI': 'TCPK nặng', 'HPrScale': 'Xỉ sơ cấp HP', 'EL': 'Lỗi xếp lớp', 'HOLE': 'Lỗ thủng',
                'RIP': 'Rách bề mặt', 'BRUS': 'Vết Hằn trục', 'LC': 'Nứt dọc', 'SCRT': 'Xước bề mặt',
                'XC': 'Xỉ cán', 'XTC': 'Xỉ thứ cấp', 'TCPK-n': 'TCPK nhẹ'
            },
            'GEOMETRY': {
                'telescope': 'Cong cạnh', 'high_spot': 'High Spot', 'dungsaitrong': 'Dung sai ĐK trong',
                'Crown': 'Độ Crown', 'Wedge': 'Độ Wedge', 'ThickDiff': 'Sai lệch dày', 'WidthDiff': 'Sai lệch rộng'
            },
            'PROPERTY': {
                'YieldPoint': 'GH Chảy', 'Tensile': 'GH Bền', 'Elongation': 'Độ giãn dài', 'Hardness': 'Độ cứng',
                'ImpactEnergy': 'Độ dai va đập',
                'C': 'Carbon', 'Mn': 'Mangan', 'Si': 'Silic', 'P': 'Photpho', 'S': 'Lưu huỳnh',
                'Cu': 'Đồng', 'Ni': 'Niken', 'Cr': 'Crom', 'Mo': 'Moly', 'V': 'Vanadi',
                'Ti': 'Titan', 'Al': 'Nhôm', 'Ca': 'Canxi', 'B': 'Bo', 'Nb': 'Niobi',
                'CEV': 'CEV', 'O': 'Oxy', 'N': 'Nitơ', 'H': 'Hydro'
            },
            'APPEARANCE': {
                'strap': 'Khuyết biên', 'label_tag': 'Bavia biên', 'packaging': 'Vỡ biên',
                'edge_cond': 'Sổ vòng', 'coil_shape': 'Loa cuộn', 'mop_bien': 'Móp biên'
            }
        }

        # 3. KHỞI TẠO FILE EXCEL
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bao_Cao_QC"

        # Định dạng styles
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        colors = {
            'INFO': '3B82F6',       # Xanh dương
            'SURFACE': 'EF4444',    # Đỏ
            'GEOMETRY': '0EA5E9',   # Xanh nhạt
            'PROPERTY': '10B981',   # Xanh lá
            'APPEARANCE': '8B5CF6', # Tím
            'NOTE': 'F59E0B'        # Cam cho cột ghi chú
        }

        # 4. CHUẨN BỊ HEADER DÒNG 1 VÀ DÒNG 2
        info_headers = [
            "STT", "ORDER", "TDC CODE", "TỔNG CẤP CHẤT LƯỢNG", "PHÂN LOẠI", 
            "TÌNH TRẠNG CHẤT LƯỢNG", "PHÂN BỔ SO", "HƯỚNG XỬ LÝ", "CHIỀU DÀY", "LỖI LỆCH TDC",
            "CN/KCN", "CA", "NGÀY SẢN XUẤT", "ID XỬ LÝ", "ID GỐC", 
            "SẢN PHẨM", "MÁC THÉP", "TIÊU CHUẨN SẢN PHẨM", "KHỐI LƯỢNG TỊNH"
        ]

        row1, row2 = [], []
        
        # Khối Thông tin chung
        for h in info_headers:
            row1.append("THÔNG TIN CHUNG")
            row2.append(h)

        # Khối Điểm Lỗi
        group_col_start = {}
        col_idx = len(info_headers) + 1

        for group_key, defect_dict in DEFECT_MAP.items():
            group_col_start[group_key] = col_idx
            for code, name in defect_dict.items():
                row1.append(group_key)
                row2.append(name)
                col_idx += 1

        # Cột Ghi chú cuối cùng
        row1.append("GHI CHÚ")
        row2.append("Ghi Chú QC")
        note_col_idx = col_idx

        ws.append(row1)
        ws.append(row2)

        # 5. GỘP Ô VÀ TÔ MÀU HEADER
        # Merge khối Thông tin chung
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(info_headers))
        for col in range(1, len(info_headers) + 1):
            c1 = ws.cell(row=1, column=col)
            c1.fill = PatternFill(start_color=colors['INFO'], end_color=colors['INFO'], fill_type="solid")
            c1.font = header_font
            c1.alignment = center_align
            c1.border = thin_border
            
            c2 = ws.cell(row=2, column=col)
            c2.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            c2.font = Font(bold=True)
            c2.alignment = center_align
            c2.border = thin_border

        # Merge các khối Nhóm Lỗi
        curr_col = len(info_headers) + 1
        for group_key, defect_dict in DEFECT_MAP.items():
            end_col = curr_col + len(defect_dict) - 1
            ws.merge_cells(start_row=1, start_column=curr_col, end_row=1, end_column=end_col)
            
            merge_cell = ws.cell(row=1, column=curr_col)
            group_title = "BỀ MẶT" if group_key == 'SURFACE' else "KÍCH THƯỚC" if group_key == 'GEOMETRY' else "CƠ/LÝ/HÓA" if group_key == 'PROPERTY' else "NGOẠI QUAN"
            merge_cell.value = group_title
            merge_cell.fill = PatternFill(start_color=colors[group_key], end_color=colors[group_key], fill_type="solid")
            merge_cell.font = header_font
            merge_cell.alignment = center_align
            merge_cell.border = thin_border

            for c in range(curr_col, end_col + 1):
                c2 = ws.cell(row=2, column=c)
                bg_color = "FEE2E2" if group_key == 'SURFACE' else "E0F2FE" if group_key == 'GEOMETRY' else "D1FAE5" if group_key == 'PROPERTY' else "EDE9FE"
                c2.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                c2.font = Font(bold=True)
                c2.alignment = center_align
                c2.border = thin_border
                
            curr_col = end_col + 1

        # Cột Ghi chú (Gộp dọc từ dòng 1 xuống dòng 2)
        ws.merge_cells(start_row=1, start_column=note_col_idx, end_row=2, end_column=note_col_idx)
        note_cell = ws.cell(row=1, column=note_col_idx)
        note_cell.value = "GHI CHÚ"
        note_cell.fill = PatternFill(start_color=colors['NOTE'], end_color=colors['NOTE'], fill_type="solid")
        note_cell.font = header_font
        note_cell.alignment = center_align
        note_cell.border = thin_border
        ws.cell(row=2, column=note_col_idx).border = thin_border

        # 6. ĐIỀN DỮ LIỆU
        for index, r in enumerate(rows):
            try:
                scores = json.loads(r['scores']) if r['scores'] else {}
            except:
                scores = {}

            # Xử lý format chuỗi SẢN PHẨM (Độ dày x Khổ rộng)
            thick = r.get('target_thick')
            width = r.get('target_width')
            try:
                sp_format = f"{float(thick):.2f}x{int(float(width))}" if thick and width else ""
            except:
                sp_format = ""

            # Xử lý format NGÀY SẢN XUẤT (Chỉ lấy ngày)
            prod_date = r.get('sl_production_date', '')
            if prod_date:
                try:
                    # Hỗ trợ xử lý cả định dạng chuỗi datetime chuẩn lẫn đối tượng datetime thực tế
                    date_str = str(prod_date).split()[0]
                    date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                    prod_date_format = date_obj.strftime('%d/%m/%Y')
                except:
                    prod_date_format = str(prod_date).split()[0]
            else:
                prod_date_format = ""

            id_xuly = r.get('ID_XuLy', '')
            coil_id = r.get('coil_id', '')
            final_id_excel = f"{id_xuly}\n(Gốc: {coil_id})" if id_xuly else coil_id

            # Logic chuyển đổi giá trị Chiều dày (is_thick_pass)
            thick_pass_val = scores.get('is_thick_pass', -1)
            if thick_pass_val == 1:
                thick_pass_str = "ĐẠT"
            elif thick_pass_val == 0:
                thick_pass_str = "Không đạt"
            else:
                thick_pass_str = ""

            # Sắp xếp đúng theo thứ tự cấu hình info_headers mới
            row_data = [
                index + 1,                          # STT
                r.get('Order', ''),                 # ORDER
                r.get('tdc_code', ''),              # TDC CODE
                r.get('quality_class', ''),         # TỔNG CẤP CHẤT LƯỢNG (Lấy quality_class thay vì quality_level)
                r.get('prime_status', ''),          # PHÂN LOẠI
                r.get('qc_status', ''),             # TÌNH TRẠNG CHẤT LƯỢNG
                r.get('mapped_po', ''),             # PHÂN BỔ SO
                r.get('rework_status', ''),         # HƯỚNG XỬ LÝ
                thick_pass_str,                     # CHIỀU DÀY
                r.get('qc_msg', ''),                # LỖI LỆCH TDC
                '',                                 # CN/KCN để trống
                r.get('Ca', ''),                    # CA
                prod_date_format,                   # NGÀY SẢN XUẤT (đã đổi nguồn sang bảng sanluong)
                r.get('ID_XuLy', ''),               # ID XỬ LÝ (Tách riêng biệt)
                r.get('coil_id', ''),               # ID GỐC (Tách riêng biệt)
                sp_format,                          # SẢN PHẨM
                r.get('grade', ''),                 # MÁC THÉP
                r.get('TieuChuan', ''),             # TIÊU CHUẨN SẢN PHẨM
                r.get('weight', '')                 # KHỐI LƯỢNG TỊNH
            ]

            # Đẩy điểm (Scores) vào row (Định dạng C1, C2...)
            for group_key, defect_dict in DEFECT_MAP.items():
                for code in defect_dict.keys():
                    val = scores.get(code, 0)
                    try:
                        val_int = int(float(val))
                        display_val = val_int if val_int > 0 else ""
                    except:
                        display_val = ""
                    row_data.append(display_val)

            # Đẩy Ghi chú vào cột cuối cùng
            raw_note = r.get('note_qc', '')
            clean_note = raw_note
            if raw_note and raw_note.startswith('{'): 
                try:
                    note_dict = json.loads(raw_note)
                    parts = [str(v) for v in note_dict.values() if v]
                    clean_note = ", ".join(parts)
                except:
                    pass
            row_data.append(clean_note)

            ws.append(row_data)
        ws.column_dimensions['N'].width = 18 # Cột ID XỬ LÝ
        ws.column_dimensions['O'].width = 18 # Cột ID GỐC
        ws.column_dimensions['P'].width = 13 # Cột SẢN PHẨM
        ws.column_dimensions['Q'].width = 12 # Cột MÁC THÉP
        ws.column_dimensions['R'].width = 15 # Cột TIÊU CHUẨN SẢN PHẨM
        ws.column_dimensions[openpyxl.utils.get_column_letter(note_col_idx)].width = 30 # Cột Ghi chú cuố

        # 7. XUẤT FILE
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        return send_file(
            out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Bao_Cao_QC.xlsx'
        )

    except Exception as e:
        import traceback
        print(f"Lỗi xuất Excel: {traceback.format_exc()}")
        return jsonify({'status': 'error', 'msg': str(e)}), 500
@dashboard_bp.route('/api/check_new_data', methods=['GET'])
def check_new_data():
    """API để Frontend polling: Hỏi xem có dữ liệu mới không"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        # Lấy thời gian dòng vừa được sửa/tạo gần nhất
        cursor.execute("SELECT MAX(updated_at) FROM coil_data WITH (NOLOCK)")
        row = cursor.fetchone()
        conn.close()
        
        if row and row[0]:
            # Chuyển đổi datetime sang timestamp
            latest_time = row[0].timestamp()
            return jsonify({'timestamp': latest_time})
    except Exception as e:
        pass
    from utils.sync_worker import LAST_DATA_UPDATE
    return jsonify(LAST_DATA_UPDATE)

@dashboard_bp.route('/qlcl', methods=['GET', 'POST'])
@permission_required('qlcl_view')
def qlcl_page():
    # NẾU LÀ POST (Nhấn nút Lọc): Lưu dữ liệu vào session rồi chuyển hướng sang GETf
    if request.method == 'POST':
        session['factory'] = request.form.get('factory', 'HRC1')
        session['grade'] = request.form.get('grade', 'ALL')
        session['Ca'] = request.form.get('Ca', 'ALL')
        session['start_date'] = request.form.get('start_date', '')
        session['end_date'] = request.form.get('end_date', '')
        session['coil_ids'] = request.form.get('coil_ids', '')
        session['order_ids'] = request.form.get('order_ids', '')
        # Chuyển hướng chính trang này nhưng bằng phương thức GET
        return redirect(url_for('dashboard_bp.qlcl_page'))
        
    # NẾU LÀ GET (F5 hoặc mới vào): Render giao diện
    return render_dashboard_logic()

@dashboard_bp.route('/api/sync_single_coil', methods=['POST'])
def sync_single_coil_endpoint():
    try:
        req = request.json
        coil_id = str(req.get('coil_id', '')).strip()
        if not coil_id: return jsonify({'status': 'error', 'msg': 'Thiếu ID'})
        
        status_report = []

        # --- BƯỚC 1: HÌNH HỌC (Geometry) ---
        geo_found = False
        try:
            # Gọi API lấy data
            resp = requests.get(f"{API_GEO_SINGLE_URL}{coil_id}", timeout=5)
            if resp.status_code == 200:
                data = resp.json()
                rows = []
                if isinstance(data, list): rows = data
                elif isinstance(data, dict): rows = data.get('data') or data.get('rows') or []
                
                if rows:
                    # Gọi hàm lưu và kiểm tra kết quả
                    updated_ids = process_and_save_geometry(rows)
                    if updated_ids: 
                        geo_found = True
        except Exception as e:
            print(f"Geo Err: {e}")

        if geo_found: status_report.append("📐 Geo: ✅")
        else: status_report.append("📐 Geo: ❌")

        # --- BƯỚC 2: BỀ MẶT (Surface) ---
        try:
            # Hàm này giờ trả về số lượng (int)
            count = sync_surface_defects([coil_id])
            if count > 0: status_report.append("🔍 Surf: ✅")
            else: status_report.append("🔍 Surf: ❌ (Không có)")
        except: 
            status_report.append("🔍 Surf: ⚠️ Lỗi")

        # --- BƯỚC 3: CƠ/LÝ/HÓA (Properties) ---
        try:
            updated_list = sync_properties_mysql([coil_id])
            
            if len(updated_list) > 0: # <--- Sửa logic check tại đây
                status_report.append("🧪 Prop: ✅")
            else: 
                status_report.append("🧪 Prop: ❌ (Không có)")
        except Exception as e: 
            status_report.append(f"🧪 Prop: ⚠️ Lỗi ({str(e)})")

        # --- TỔNG KẾT ---
        final_msg = f"Kết quả quét {coil_id}: <br/>" + " | ".join(status_report)
        
        return jsonify({'status': 'success', 'msg': final_msg})

    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})

@dashboard_bp.route('/api/sync_batch_coils', methods=['POST'])
def sync_batch_coils_endpoint():
    try:
        req = request.json
        raw_ids = req.get('coil_ids', '')
        # Tách và làm sạch ID
        coil_ids = [x.strip().upper() for x in raw_ids.replace('\n', ',').split(',') if x.strip()]
        
        if not coil_ids: return jsonify({'status': 'error', 'msg': 'Danh sách ID trống!'})

        # --- HÀM CHẠY NGẦM ---
        def run_batch_job(target_ids):
            print(f"🚀 [Batch Scan] Đang xử lý {len(target_ids)} cuộn...")
            prefix_hrc1 = ('7', 'A', 'C', 'D')
            # --- BƯỚC 1: PHÂN LOẠI NHÀ MÁY (Dựa trên ký tự đầu) ---
            # HRC1: Bắt đầu bằng 7
            ids_hrc1 = [cid for cid in target_ids if cid.startswith(prefix_hrc1)]
            # HRC2: Bắt đầu bằng 8
            ids_hrc2 = [cid for cid in target_ids if not cid.startswith(prefix_hrc1)]

            # --- BƯỚC 2: QUÉT CƠ TÍNH / HÓA HỌC (MySQL) ---
            # Gọi riêng cho từng nhà máy để trỏ đúng View SQL
            if ids_hrc1:
                try:
                    print(f"🧪 [MySQL] Quét {len(ids_hrc1)} cuộn bên HRC1...")
                    sync_properties_mysql(ids_hrc1, factory_id='HRC1')
                except Exception as e:
                    print(f"❌ [MySQL HRC1 Error] {e}")

            if ids_hrc2:
                try:
                    print(f"🧪 [MySQL] Quét {len(ids_hrc2)} cuộn bên HRC2...")
                    sync_properties_mysql(ids_hrc2, factory_id='HRC2')
                except Exception as e:
                    print(f"❌ [MySQL HRC2 Error] {e}")

            # --- BƯỚC 3: QUÉT API (SURFACE & GEOMETRY) ---
            # Lặp qua toàn bộ danh sách tổng để chạy API
            for cid in target_ids:
                try:
                    # Định tuyến lại lần nữa cho chắc chắn
                    current_fid = 'HRC1' if cid.startswith(prefix_hrc1) else 'HRC2'
                    
                    cfg = FACTORY_CONFIGS.get(current_fid)
                    if not cfg: continue

                    # A. Quét Bề mặt TRƯỚC (Lấy lỗi, chấp nhận metadata 0)
                    sync_surface_defects([cid], factory_id=current_fid)
                    
                    # B. Quét Hình học SAU CÙNG (Lấy metadata weight/date/thick đè lên)
                    geo_url = f"{cfg['api_geo']}{cid}"
                    geo_resp = requests.get(geo_url, timeout=3)
                    
                    if geo_resp.status_code == 200:
                        data = geo_resp.json()
                        rows = []
                        if isinstance(data, list): rows = data
                        elif isinstance(data, dict): rows = data.get('data', [])
                        
                        if rows:
                            # Quan trọng: Truyền factory_id để lưu đúng cột factory
                            process_and_save_geometry(rows, factory_id=current_fid)
                    
                    time.sleep(1) # Tránh Dos server

                except Exception as e:
                    print(f"⚠️ Lỗi API cuộn {cid}: {e}")

            print(f"🏁 [Batch Scan] Hoàn tất toàn bộ.")

        # Khởi động thread
        thread = threading.Thread(target=run_batch_job, args=(coil_ids,))
        thread.start()

        return jsonify({'status': 'success', 'msg': f'Hệ thống đang xử lý {len(coil_ids)} cuộn (Phân luồng HRC1/HRC2 tự động).'})

    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
## RENDER HEATMAP 
def get_menu_structure_for_grade(grade_config):
    menu = {'surface': [], 'geometry': [], 'prop': []}
    if not grade_config: return menu
    for name, cfg in grade_config.items():
        grp = cfg.get('group')
        if grp == 'surface': menu['surface'].append(name)
        elif grp == 'geometry': menu['geometry'].append(name)
        elif grp in ['mechanical', 'chemical']: menu['prop'].append(name)
    return menu

def render_dashboard_logic(msg=None):
    t_start_total = time.time()
    current_factory = request.args.get('factory') or session.get('factory', 'HRC1')
    selected_grade = request.args.get('grade') or session.get('grade', 'ALL')
    selected_ca = request.args.get('Ca') or session.get('Ca', 'ALL')
    start_date = request.args.get('start_date') or session.get('start_date', '')
    end_date = request.args.get('end_date') or session.get('end_date', '')
    raw_coil_ids = request.args.get('coil_ids') or session.get('coil_ids', '')
    raw_order_ids = request.args.get('order_ids') or session.get('order_ids', '')
    # Lưu ngược lại session để F5 luôn nhớ trạng thái cuối cùng
    session['factory'] = current_factory
    session['grade'] = selected_grade
    if not start_date and not end_date:
        today_dt = datetime.datetime.now()
        start_date = (today_dt - datetime.timedelta(days=1)).strftime('%Y-%m-%dT00:00')
        end_date = today_dt.strftime('%Y-%m-%dT23:59')
    elif not raw_coil_ids and start_date and end_date:
        try:
            fmt_start = '%Y-%m-%dT%H:%M' if 'T' in start_date else '%Y-%m-%d'
            fmt_end = '%Y-%m-%dT%H:%M' if 'T' in end_date else '%Y-%m-%d'
            d_start = datetime.datetime.strptime(start_date, fmt_start)
            d_end = datetime.datetime.strptime(end_date, fmt_end)
            if (d_end - d_start).days > 7:
                d_end = d_start + datetime.timedelta(days=7)
                end_date = d_end.strftime('%Y-%m-%dT%H:%M')
                msg = "Giao diện Dashboard chỉ hiển thị tối đa 7 ngày. Dữ liệu đã được tự động điều chỉnh."
        except Exception:
            pass
    session['start_date'] = start_date
    session['end_date'] = end_date
    session['coil_ids'] = raw_coil_ids
    session['order_ids'] = raw_order_ids
    conn = None
    rows = []
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        sql = """
            SELECT c.coil_id, c.ID_XuLy, c.grade, c.raw_data, c.scores, c.is_checked, 
                   c.production_date, c.updated_at, c.factory ,
                   c.Temperature, c.Speed, c.quality_level, c.note_qc,
                   c.slab_grade, c.weight, c.target_thick, c.target_width , c.Nhom,
                   c.qc_stage, c.qc_status, c.mapped_po, c.rework_status,
                   c.suggested_order_map, c.qc_msg, c.quality_class, c.prime_status,
                   c.stage1_penalty, c.stage2_penalty,
                   c.downgrade_reason,
                   ISNULL(r.is_skin_required, 0) as is_skin_required,
                   c.[Order] as original_order,
                   tm.tdc_code,
                   c.TARGET_LV2,
                   tv.criteria_json
            FROM coil_data c WITH (NOLOCK) 
            LEFT JOIN order_production_rules r WITH (NOLOCK) ON c.[Order] = r.[Order]
            -- BỔ SUNG JOIN LẤY TDC
            LEFT JOIN tdc_versions tv WITH (NOLOCK) ON c.target_tdc_version_id = tv.id
            LEFT JOIN tdc_master tm WITH (NOLOCK) ON tv.master_id = tm.id
            WHERE 1=1 
        """
        params = []
        coil_id_list = []
        order_id_list = []
        if raw_coil_ids:
            # 1. Băm chuỗi: \s sẽ tóm gọn toàn bộ dấu cách, tab, và xuống dòng (\n)
            raw_list = re.split(r'[\s,;]+', raw_coil_ids)
            # 2. Làm sạch: Xóa khoảng trắng thừa từng ID, in hoa, và bỏ ID rỗng
            clean_list = [x.strip().upper() for x in raw_list if x.strip()]
            coil_id_list = list(dict.fromkeys(clean_list))
        if raw_order_ids:
            raw_list_ord = re.split(r'[\s,;]+', raw_order_ids)
            order_id_list = list(dict.fromkeys([x.strip().upper() for x in raw_list_ord if x.strip()]))
        # LOGIC ƯU TIÊN LỌC ID
        if coil_id_list:
            if len(coil_id_list) > 1000:
                coil_id_list = coil_id_list[:1000]
            placeholders = ', '.join(['?'] * len(coil_id_list))
            sql += f" AND (coil_id IN ({placeholders}) OR ID_XuLy IN ({placeholders}))"
            params.extend(coil_id_list)
            params.extend(coil_id_list)
            selected_grade = 'ALL' 
        elif order_id_list:
            # ƯU TIÊN 2: LỌC THEO DANH SÁCH ORDER
            if len(order_id_list) > 1000: order_id_list = order_id_list[:1000]
            placeholders = ', '.join(['?'] * len(order_id_list))
            sql += f" AND c.[Order] IN ({placeholders})"
            params.extend(order_id_list)
            selected_grade = 'ALL'
        else:
            sql += " AND c.factory = ?"
            params.append(current_factory)

            if selected_grade != 'ALL':
                sql += " AND c.grade = ?"
                params.append(selected_grade)
            if selected_ca != 'ALL':
                sql += " AND c.Ca = ?"
                params.append(selected_ca)
            if start_date:
                s_str = start_date.replace('T', ' ')
                if len(s_str) == 10: s_str += " 00:00:00" 
                elif len(s_str) == 16: s_str += ":00"      
                sql += " AND c.production_date >= ?"
                params.append(s_str)
            if end_date:
                e_str = end_date.replace('T', ' ')
                if len(e_str) == 10: e_str += " 23:59:59.999"
                elif len(e_str) == 16: e_str += ":59.999"
                sql += " AND c.production_date <= ?"
                params.append(e_str)

        sql += " ORDER BY c.production_date DESC, c.coil_id DESC"
        
        cursor.execute(sql, tuple(params))
        rows = db.fetchall_as_dict(cursor) 
    except Exception as e:
        print(f"Lỗi lấy dữ liệu Dashboard: {e}")
    finally:
        if conn: conn.close()
    all_data = {
        r['coil_id']: {
            'scores': json.loads(r['scores']) if r['scores'] else {}, 
            'raw_data': json.loads(r['raw_data']) if r['raw_data'] else {}, 
            'GRADE': r['grade'], 
            'IS_CHECKED': r['is_checked'],
            'updated_at': r['updated_at'],
            'production_date': str(r['production_date']) if r['production_date'] else '',
            'Temperature': r['Temperature'] if r['Temperature'] else 0,
            'Speed': r['Speed'] if r['Speed'] else 0,
            'quality_level': r['quality_level'] if r['quality_level'] else '',
            'note_qc': r['note_qc'] if r['note_qc'] else '',
            'slab_grade': r['slab_grade'] if r['slab_grade'] else '---',
            'weight': float(r['weight']) if r['weight'] else 0.0,
            'target_thick': float(r['target_thick']) if r['target_thick'] else 0,
            'target_width': float(r['target_width']) if r['target_width'] else 0,
            'ID_xuly': r['ID_XuLy'] if 'ID_XuLy' in r and r['ID_XuLy'] else '',
            'Nhom': r['Nhom'] if 'Nhom' in r and r['Nhom'] else '',
            'qc_stage': r['qc_stage'] if 'qc_stage' in r and r['qc_stage'] else '',
            'qc_status': r['qc_status'] if 'qc_status' in r and r['qc_status'] else '',
            'mapped_po': r['mapped_po'] if 'mapped_po' in r and r['mapped_po'] else '',
            'rework_status': r['rework_status'] if 'rework_status' in r and r['rework_status'] else 'NULL',
            'suggested_order_map': r['suggested_order_map'] if 'suggested_order_map' in r and r['suggested_order_map'] else '', # [THÊM MỚI]
            'qc_msg': r['qc_msg'] if 'qc_msg' in r and r['qc_msg'] else '',
            'quality_class': r['quality_class'] if 'quality_class' in r and r['quality_class'] else '',
            'prime_status': r['prime_status'] if 'prime_status' in r and r['prime_status'] else '',
            'downgrade_reason': r['downgrade_reason'] if 'downgrade_reason' in r and r['downgrade_reason'] else '',
            'is_skin_required': r['is_skin_required'] if 'is_skin_required' in r else 0, 
            'original_order': r['original_order'] if 'original_order' in r and r['original_order'] else '---',
            'tdc_code': r['tdc_code'] if 'tdc_code' in r and r['tdc_code'] else '---',
            'TARGET_LV2': float(r['TARGET_LV2'] or 0.0),
            'criteria_json': r.get('criteria_json'),
            'stage1_penalty': r.get('stage1_penalty', 0), 
            'stage2_penalty': r.get('stage2_penalty', 0),
        } 
        for r in rows
    }
    dashboard_data = regenerate_dashboard_data(all_data, selected_grade)
    
    data_wrapper = {
        'has_data': bool(all_data),
        'time_range': db.get_config('time_range', ''),
        'radar_data': dashboard_data['radar_data'],
        'tabs': dashboard_data['tabs'],
    }
    
    t_end_total = time.time()
    try:
        with open('perf_log.txt', 'a', encoding='utf-8') as f:
            f.write(f"[{datetime.datetime.now()}] render_dashboard_logic Total: {t_end_total - t_start_total:.4f}s\n")
    except: pass

    return render_template('qlcl.html', 
            data=data_wrapper, 
            radar_data=data_wrapper['radar_data'], 
            menu=dashboard_data['menu'], 
            current_grade=selected_grade,
            current_factory=current_factory,
            current_ca=selected_ca,
            start_date=start_date,
            end_date=end_date,
            coil_ids=raw_coil_ids,
            order_ids=raw_order_ids,
            msg=msg
        )
def regenerate_dashboard_data(all_data, selected_grade):
    all_configs = get_all_grade_configs()
    current_config = all_configs.get(selected_grade)
    if not current_config:
        current_config = all_configs.get('SAE1006')
    final_radar_data = {} # Dữ liệu gửi xuống Frontend (Chứa TOÀN BỘ cuộn)
    target_coils = []     # Dữ liệu để tính toán Tabs thống kê (Chỉ Mác đang chọn)

    # 1. DUYỆT QUA TẤT CẢ CUỘN (Không lọc ngay đầu để tránh mất dữ liệu tìm kiếm)
    for cid, d in all_data.items():
        raw = d.get('raw_data', {})
        current_scores = d.get('scores', {})
        db_grade = d.get('GRADE') if d.get('GRADE') else 'SAE1006'
        db_grade_clean = str(db_grade).strip().upper()

        if raw:
             thick_val = float(d.get('TARGET_LV2') or 0.0)
             auto_scores = process_coil_scores(cid, raw, db_grade_clean, thickness=thick_val, cached_config=all_configs)
        else:
             auto_scores = {}
        raw_crit = d.get('criteria_json')
        tdc_limits = {}
        if raw_crit:
            try:
                crit_list = json.loads(raw_crit) if isinstance(raw_crit, str) else raw_crit
                for crit in crit_list:
                    defect = crit.get('defect')
                    allowed_range = crit.get('range', [])
                    if defect and allowed_range:
                        tdc_limits[defect] = {
                            'min': min(allowed_range),
                            'max': max(allowed_range)
                        }
            except: pass
        frontend_obj = current_scores.copy() 
        frontend_obj['tdc_limits'] = tdc_limits
        frontend_obj['auto_scores'] = auto_scores
        optimized_raw = {}
        if raw:
            for k, v in raw.items():
                # --- ĐOẠN NÀY LÀ CODE MỚI CẦN THÊM VÀO ---
                grade_config = all_configs.get(db_grade_clean, all_configs.get('SAE1006'))

                if isinstance(v, list):
                    cfg = grade_config.get(k, {})
                    optimized_raw[k] = build_matrix_tooltip(v, cfg)
                else:
                    optimized_raw[k] = v
        
        frontend_obj['raw_data'] = optimized_raw
        frontend_obj['GRADE'] = db_grade_clean
        frontend_obj['IS_CHECKED'] = d.get('IS_CHECKED', False)
        prod_date = d.get('production_date', '')
        frontend_obj['production_date'] = str(prod_date) if prod_date else ''
        frontend_obj['Temperature'] = d.get('Temperature', 0)
        frontend_obj['Speed'] = d.get('Speed', 0)
        frontend_obj['quality_level'] = d.get('quality_level', '')
        frontend_obj['note_qc'] = d.get('note_qc', '')
        frontend_obj['slab_grade'] = d.get('slab_grade', '---')
        frontend_obj['weight'] = d.get('weight', 0.0)
        frontend_obj['target_thick'] = d.get('target_thick', 0)
        frontend_obj['target_width'] = d.get('target_width', 0)
        frontend_obj['ID_xuly'] = d.get('ID_xuly', '')
        frontend_obj['Nhom'] = d.get('Nhom', '')
        frontend_obj['qc_stage'] = d.get('qc_stage', '')
        frontend_obj['qc_status'] = d.get('qc_status', '')
        frontend_obj['mapped_po'] = d.get('mapped_po', '')
        frontend_obj['rework_status'] = d.get('rework_status', 'NULL')
        frontend_obj['suggested_order_map'] = d.get('suggested_order_map', '')
        frontend_obj['qc_msg'] = d.get('qc_msg', '') 
        frontend_obj['quality_class'] = d.get('quality_class', '')  
        frontend_obj['prime_status'] = d.get('prime_status', '')
        frontend_obj['downgrade_reason'] = d.get('downgrade_reason', '')
        frontend_obj['is_skin_required'] = d.get('is_skin_required', 0)
        frontend_obj['original_order'] = d.get('original_order', '---')
        frontend_obj['tdc_code'] = d.get('tdc_code', '---')
        frontend_obj['stage1_penalty'] = d.get('stage1_penalty', 0) 
        frontend_obj['stage2_penalty'] = d.get('stage2_penalty', 0)
        final_radar_data[cid] = frontend_obj
        if selected_grade == 'ALL' or db_grade_clean == selected_grade:
            target_coils.append({
                'CustomerID': cid, 
                'Raw': raw,
                'Scores': current_scores,
                'Grade': db_grade_clean 
            })
    
    # --- Performance Logging Start ---
    t_start_regen = time.time()
    
    if not target_coils:
        return {
            'tabs': {}, 
            'radar_data': final_radar_data, 
            'menu': get_menu_structure_for_grade(current_config)
        }

    score_lookup = {item['CustomerID']: item['Scores'] for item in target_coils}

    # --- VECTORIZED OPTIMIZATION START ---
    t_start_pandas = time.time()
    
    # 1. Create Main DataFrame directly from target_coils
    # Use list comprehension for metadata and pd.DataFrame for Raw dicts (faster than json_normalize)
    df_meta = pd.DataFrame(target_coils, columns=['CustomerID', 'Grade'])
    df_raw = pd.DataFrame([item['Raw'] for item in target_coils])
    df_main = pd.concat([df_meta, df_raw], axis=1)

    # 2. Prepare Surface DataFrame (df_surface)
    # Identify Surface Columns based on Current Config
    surface_keys = [
        k for k, cfg in current_config.items() 
        if cfg.get('group') == 'surface' or cfg.get('mode') in ['count', 'matrix']
    ]
    
    # Only process columns that actually exist in data
    valid_surface_keys = [k for k in surface_keys if k in df_main.columns]

    if valid_surface_keys:
        # A. Filter & Melt
        df_surf_subset = df_main[['CustomerID'] + valid_surface_keys].copy()
        df_melt = df_surf_subset.melt(id_vars=['CustomerID'], value_vars=valid_surface_keys, var_name='DefectClass', value_name='Size')
        
        # B. Handle "Count" logic (Int -> List[0,0...]) vs "Matrix" logic (List of Sizes)
        # Vectorized apply is faster than explicit loop
        def normalize_size_vectorized(val):
            if isinstance(val, list):
                return val
            # Handle scalar counts (e.g. 5 -> [0,0,0,0,0])
            try:
                if pd.notna(val) and val != '':
                    v_int = int(val)
                    if v_int > 0: return [0] * v_int
            except: pass
            return []

        df_melt['Size'] = df_melt['Size'].apply(normalize_size_vectorized)
        
        # C. Explode & Clean
        df_surface = df_melt.explode('Size')
        # Remove rows with empty/NaN Size (exploded from empty list or None)
        df_surface = df_surface.dropna(subset=['Size'])
    else:
        df_surface = pd.DataFrame(columns=['CustomerID', 'DefectClass', 'Size'])

    t_end_pandas = time.time()
    # --- VECTORIZED OPTIMIZATION END ---

    tabs_data = {}
    total_rolls = len(target_coils)
    
    t_start_metrics = time.time()
    for name, cfg in current_config.items():
        if cfg.get('group') == 'surface':
            tabs_data[name] = calculate_metric_surface(df_surface, name, cfg, total_rolls)
        else:
            tabs_data[name] = calculate_metric_value(df_main, name, cfg, total_rolls, score_lookup)
    t_end_metrics = time.time()

    # Log Performance
    try:
        with open('perf_log.txt', 'a', encoding='utf-8') as f:
            f.write(f"[{datetime.datetime.now()}] Vectorized Regenerate:\n")
            f.write(f"  > DataFrame Prep: {t_end_pandas - t_start_pandas:.4f}s\n")
            f.write(f"  > Metrics Calc: {t_end_metrics - t_start_metrics:.4f}s\n")
    except: pass

    return {'tabs': tabs_data, 'radar_data': final_radar_data, 'menu': get_menu_structure_for_grade(current_config)}

@dashboard_bp.route('/api/recalc_scores_by_grade', methods=['POST'])
def recalc_scores_by_grade():
    try:
        req = request.json
        target_grade = req.get('grade')
        
        if not target_grade: return jsonify({'status': 'error', 'msg': 'Thiếu tên Mác thép'})

        conn = db.get_connection()
        cursor = conn.cursor()
        query = """
            SELECT coil_id, raw_data, grade, scores, is_checked, factory,
                production_date, slab_grade, weight, target_thick, target_width, note_qc ,quality_level, Temperature, Speed, TARGET_LV2
            FROM coil_data WITH (NOLOCK) 
            WHERE production_date >= '2026-07-01 08:00:00'  and grade = ?
        """
        cursor.execute(query, (target_grade,))
        rows = db.fetchall_as_dict(cursor)
        conn.close()

        if not rows:
            return jsonify({'status': 'success', 'msg': f'Không tìm thấy cuộn nào thuộc mác {target_grade}.'})

        count_total = len(rows)

        # ====================================================
        # HÀM CHẠY NGẦM ĐỂ KHÔNG LÀM TREO GIAO DIỆN
        # ====================================================
        def run_recalc_background(coils_to_process):
            import time
            print(f"🔄 [Recalc] Đang tính lại điểm cho {count_total} cuộn mác {target_grade}...")
            fresh_configs = get_all_grade_configs()
            chunk_size = 500  # CHIA LÔ: Mỗi lần chỉ xử lý và lưu 500 cuộn
            total_processed = 0

            for i in range(0, len(coils_to_process), chunk_size):
                chunk = coils_to_process[i:i + chunk_size]
                batch_update = []

                for r in chunk:
                    coil_id = r['coil_id']
                    raw = json.loads(r['raw_data']) if r['raw_data'] else {}
                    old_scores = json.loads(r['scores']) if r['scores'] else {}
                    
                    thick_val = float(r.get('TARGET_LV2') or 0.0)
                    new_scores = process_coil_scores(coil_id, raw, target_grade, thickness=thick_val, cached_config=fresh_configs)
                    
                    final_scores = new_scores
                    if r['is_checked'] == 1:
                        final_scores = old_scores.copy()
                        
                        # Khai báo danh sách các chất cần ép buộc tính lại
                        MECH_KEYS = [
                            'YieldPoint', 'Tensile', 'Elongation', 'Hardness',
                            'C', 'Mn', 'Si', 'P', 'S', 'Cu', 'Ni', 'Cr', 'Mo', 
                            'V', 'Ti', 'Al', 'Ca', 'B', 'Nb', 'CEV', 'O', 'N', 'H','Crown','WidthDiff'
                        ]
                        
                        for k, v in new_scores.items():
                            # 1. Nếu là Cơ tính/Hóa học: LUÔN CẬP NHẬT theo Rule mới nhất
                            if k in MECH_KEYS:
                                final_scores[k] = v
                            
                            # 2. Nếu là Bề mặt/Hình học: CHỈ VÁ khi điểm cũ đang trống (bằng 0)
                            elif old_scores.get(k, 0) == 0:
                                final_scores[k] = v
                                
                    batch_update.append({
                        'id': coil_id, 'grade': target_grade, 'raw': raw, 'scores': final_scores, 
                        'is_checked': r['is_checked'], 'factory': r.get('factory', 'HRC1'),
                        'production_date': r.get('production_date'), 'slab_grade': r.get('slab_grade'),
                        'weight': r.get('weight', 0), 'target_thick': r.get('target_thick', 0),
                        'target_width': r.get('target_width', 0), 'Temperature': r.get('Temperature', 0),
                        'Speed': r.get('Speed', 0), 'quality_level': r.get('quality_level'), 'note_qc': r.get('note_qc'),'TARGET_LV2': r.get('TARGET_LV2', 0)
                    })

                # Lưu 1 lô 500 cuộn
                if batch_update:
                    db.save_batch_coils_v2(batch_update)
                    total_processed += len(batch_update)
                    print(f"   -> Đã tính xong {total_processed}/{count_total} cuộn.")
                    
                time.sleep(0.5) # NGHỈ 0.5 GIÂY: Nhường tài nguyên CPU & DB cho luồng dây chuyền (Real-time)

            print(f"🏁 [Recalc] Đã hoàn tất tính lại điểm cho mác {target_grade}!")
            try:
                from routes.dashboard import log_system_event
                log_system_event("Tính lại điểm", f"Đã cập nhật điểm số cho {count_total} cuộn mác {target_grade}.", "success")
            except: pass

        # Kích hoạt luồng chạy ngầm
        import threading
        threading.Thread(target=run_recalc_background, args=(rows,)).start()

        # Trả về ngay lập tức cho Web (Không bị Timeout)
        return jsonify({
            'status': 'success', 
            'msg': f'Hệ thống đang chạy ngầm để tính lại điểm cho {count_total} cuộn {target_grade}. Quá trình này sẽ mất khoảng vài chục giây!'
        })

    except Exception as e:
        print(f"Recalc Error: {e}")
        return jsonify({'status': 'error', 'msg': str(e)})
@dashboard_bp.route('/api/get_input_detail/<coil_id>', methods=['GET'])
def get_input_detail(coil_id):
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # 1. Lấy dữ liệu từ DB
        cursor.execute("SELECT raw_data, scores, grade, is_checked , quality_level, note_qc  FROM coil_data WITH (NOLOCK) WHERE coil_id = ?", (coil_id,))
        row = cursor.fetchone()
        conn.close()

        if not row:
            return jsonify({'status': 'error', 'msg': 'Không tìm thấy cuộn'})

        # 2. Parse dữ liệu
        current_raw = json.loads(row[0]) if row[0] else {}
        current_scores = json.loads(row[1]) if row[1] else {}
        grade = row[2] if row[2] else 'SAE1006'
        
        # 3. TÍNH TOÁN DỮ LIỆU GỐC (REFERENCE) TẠI CHỖ
        all_configs = get_all_grade_configs() # Hàm này đã tối ưu cache ở câu trả lời trước
        auto_scores = process_coil_scores(coil_id, current_raw, grade, cached_config=all_configs)

        return jsonify({
            'status': 'success',
            'data': {
                'coil_id': coil_id,
                'grade': grade,
                'current_scores': current_scores, 
                'original_scores': auto_scores,   
                'raw_data': current_raw,
                'is_checked': row[3],
                'quality_level': row[4] if row and len(row)>4 else '',
                'note_qc': row[5] if row and len(row)>5 else ''
            }
        })
    except Exception as e:
        print(f"Err detail: {e}")
        return jsonify({'status': 'error', 'msg': str(e)})
    
## Nhập tay
@dashboard_bp.route('/get_manual_config', methods=['GET'])
def get_manual_config():
    return jsonify({
        'SURFACE_MANUAL': [
            {'id': 'oil', 'label': 'Gấp nếp'}, 
            {'id': 'rust', 'label': 'Nếp Nhăn'}, 
            {'id': 'scratch_m', 'label': 'Vết Hằn'}, 
            {'id': 'dirt', 'label': 'Gãy mặt'}, 
            {'id': 'mark', 'label': 'Xỉ thứ cấp'}, 
            {'id': 'scale', 'label': 'Xỉ cán'}, 
            {'id': 'other_s', 'label': 'Xỉ muối tiêu'},
            {'id': 'gianbien', 'label': 'Giãn biên/Bụng'}, 
            {'id': 'chambi', 'label': 'Chấm bi'}            
        ],
        'GEO_MANUAL': [{'id': 'telescope', 'label': 'Cong cạnh'},{'id': 'high_spot', 'label': 'High Spot'},{'id': 'dungsaitrong', 'label': 'Dung sai ĐK trong'}, ],
        'APPEARANCE': [
            {'id': 'strap', 'label': 'Khuyết biên'}, 
            {'id': 'label_tag', 'label': 'Bava biên'}, 
            {'id': 'packaging', 'label': 'Vỡ biên'}, 
            {'id': 'edge_cond', 'label': 'Sổ vòng'}, 
            {'id': 'coil_shape', 'label': 'Loa cuộn'},
            {'id': 'mop_bien', 'label': 'Móp biên'}  
        ]
    })
@dashboard_bp.route('/save_manual_data', methods=['POST'])
@permission_required('qlcl_input')
def save_manual_data():
    conn = None
    try:
        req = request.json
        coil_id = req.get('coil_id')
        new_scores = req.get('scores')
        user_name = session.get('username', 'Unknown')
        is_reset = req.get('is_reset', False)
        new_notes_dict = req.get('note_dict')
        
        if not coil_id: 
            return jsonify({'status':'error', 'msg': 'Thiếu ID cuộn'})

        conn = db.get_connection()
        cursor = conn.cursor() 
        
        # ==========================================
        # 1. LẤY DỮ LIỆU CŨ VÀ KHÓA DÒNG NÀY LẠI
        # ==========================================
        cursor.execute("""
            SELECT c.scores, c.note_qc, c.qc_stage, c.stage2_penalty, c.stage2_msg, 
                c.weight, ISNULL(c.req_min_w, 0), ISNULL(c.req_max_w, 0),
                v.criteria_json, r.[Order] as order_id, r.production_status, r.SO_mapping,
                c.qc_status, c.quality_class, c.mapped_po, c.prime_status,
                ISNULL(r.is_skin_required, 0) as is_skin_required, c.ID_XuLy, c.rework_status 
            FROM coil_data c WITH (UPDLOCK, ROWLOCK)
            LEFT JOIN tdc_versions v WITH (NOLOCK) ON c.target_tdc_version_id = v.id
            LEFT JOIN order_production_rules r WITH (NOLOCK) ON c.[Order] = r.[Order]
            WHERE c.coil_id = ?
        """, (coil_id,))

        curr_row = cursor.fetchone()
        if not curr_row:
            return jsonify({'status': 'error', 'msg': 'Không tìm thấy cuộn trong Database'})

        # Bóc tách biến (Thêm is_skin_required ở cuối)
        (old_scores_raw, old_note_raw, qc_stage, stage2_penalty, stage2_msg, 
        coil_weight, req_min_w, req_max_w, criteria_json, 
        order_id, prod_status, so_mapping, old_qc_status, old_q_class, old_mapped_po, old_p_status, is_skin_required, id_xuly, old_rework_status) = curr_row
        old_scores = json.loads(old_scores_raw) if old_scores_raw else {}
        stage2_penalty = stage2_penalty or 0
        stage2_msg = stage2_msg or ""

        # ==========================================
        # 2. XỬ LÝ GHI CHÚ (NOTE) VÀ GHI LOG (AUDIT)
        # ==========================================
        final_note_str = old_note_raw
        if new_notes_dict is not None:
            old_note_obj = {}
            if old_note_raw:
                try: old_note_obj = json.loads(old_note_raw)
                except: old_note_obj = {"app": str(old_note_raw)}
            for k, v in new_notes_dict.items(): 
                old_note_obj[k] = v
            final_note_str = json.dumps(old_note_obj, ensure_ascii=False)
        else:
            if 'note_qc' in req: final_note_str = req['note_qc']

        import datetime
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        logs = []
        for key, new_val in new_scores.items():
            old_val = old_scores.get(key, 0)
            if float(new_val) != float(old_val):
                logs.append((coil_id, user_name, key, float(old_val), float(new_val), now))

        if logs:
            cursor.executemany("INSERT INTO audit_log_qlcl (coil_id, user_name, defect_key, old_value, new_value, changed_at) VALUES (?, ?, ?, ?, ?, ?)", logs)

        # Trộn điểm cũ và điểm mới
        final_scores = old_scores.copy()
        if is_reset:
            manual_keys = ['oil', 'rust', 'scratch_m', 'dirt', 'other_s', 'gianbien','chambi','dungsaitrong', 'telescope', 'high_spot','is_thick_pass']
            for k in manual_keys:
                if k in final_scores:
                    final_scores[k] = 1
        clean_new_scores = {}
        for key, val in new_scores.items():
            if key == 'is_thick_pass' and val == -1 and old_scores.get('is_thick_pass') in [0, 1] and not is_reset:
                continue
            clean_new_scores[key] = val

        final_scores.update(clean_new_scores)
        new_is_checked = 0 if is_reset else 1
        scores_json_str = json.dumps(final_scores)

        # ==========================================
        # 🛡️ GIẢI PHÁP SỬA LỖI: THẢ LỎNG CHO PHÉP LƯU DỮ LIỆU THÔ
        # Nếu cuộn mồ côi (chưa có TDC hoặc Order), vẫn cho lưu điểm/note bình thường
        # ==========================================
        if not criteria_json or not order_id:
            cursor.execute("""
                UPDATE coil_data 
                SET scores = ?, is_checked = ?, note_qc = ?, updated_at = GETDATE()
                WHERE coil_id = ?
            """, (scores_json_str, new_is_checked, final_note_str, coil_id))
            conn.commit()
            rework_out = old_rework_status
            return jsonify({
                'status': 'success', 
                'msg': '✅ Đã lưu điểm và ghi chú thành công! (Cuộn chưa gán Đơn hàng/TDC nên hệ thống tạm bỏ qua luồng đánh giá tự động)',
                'qc_status': qc_stage or 'NULL',
                'rework_status': rework_out
            })

        # ==========================================
        # 3. KÍCH HOẠT LOGIC CASE 5: ĐÁNH GIÁ CHẤT LƯỢNG TỰ ĐỘNG
        # ==========================================
        from utils.scoring import evaluate_tdc_stage_1, evaluate_tdc_stage_2

        # 🌟 1. TỰ ĐỘNG NỘI SUY STAGE (Giống điều kiện ETL)
        try:
            has_mech = (
                float(final_scores.get('YieldPoint', 0) or 0) > 0 and
                float(final_scores.get('Tensile', 0) or 0) > 0 and
                float(final_scores.get('Elongation', 0) or 0) > 0
            )
        except Exception:
            has_mech = False

        actual_qc_stage = 'STAGE_2' if has_mech else 'STAGE_1'

        # 🌟 2. CHẠY ĐÁNH GIÁ STAGE 1 (Bắt buộc)
        res1 = evaluate_tdc_stage_1(scores_json_str, criteria_json, coil_weight, req_min_w, req_max_w)
        new_stage1_penalty = res1['stage1_penalty']
        new_stage1_msg = res1['stage1_msg']

        # 🌟 3. CHẠY ĐÁNH GIÁ STAGE 2
        if actual_qc_stage == 'STAGE_2':
            res2 = evaluate_tdc_stage_2(scores_json_str, criteria_json)
            new_stage2_penalty = res2['stage2_penalty']
            new_stage2_msg = res2['stage2_msg']
        else:
            # FIX LỖI: Chặn rác nếu bị thụt lùi từ Stage 2 về Stage 1
            new_stage2_penalty = 0
            new_stage2_msg = ""

        total_penalty = new_stage1_penalty + new_stage2_penalty
        is_downgraded = old_p_status in ['NON_PRIME', 'SCRAP']
        
        # KHỞI TẠO TRƯỚC ĐỂ TRÁNH LỖI UNBOUND LOCAL ERROR
        final_mapped_po = old_mapped_po 
        
        # -----------------------------------------------------------------
        # A. XÁC ĐỊNH TRẠNG THÁI
        # -----------------------------------------------------------------
        if total_penalty == 0:
            final_status = 'PASS' if actual_qc_stage == 'STAGE_2' else 'PASSNOCHEM'
            final_msg = ""
            final_q_class = old_q_class if is_downgraded else 'LOAI_1'
            final_p_status = old_p_status if is_downgraded else 'PRIME'
            # final_mapped_po giữ nguyên old_mapped_po
        else:
            final_status = 'FAIL' if actual_qc_stage == 'STAGE_2' else 'FAILNOCHEM'
            msgs = [m for m in [new_stage1_msg, new_stage2_msg] if m and m != "Đạt"]
            final_msg = " | ".join(msgs)
            
            if is_downgraded:
                final_status = 'PASS' # Bảo toàn trạng thái cho hàng rớt đã xử lý KCS
                final_q_class = old_q_class
                final_p_status = old_p_status
                # final_mapped_po giữ nguyên old_mapped_po
            else:
                final_q_class = None
                final_p_status = None
                final_mapped_po = '0' # Rớt thì mất quyền giữ cờ PO

        # -----------------------------------------------------------------
        # B. TÍNH TOÁN CỘNG TRỪ SẢN LƯỢNG
        # -----------------------------------------------------------------
        safe_old_q_class = old_q_class if old_q_class else 'LOAI_1'
        safe_final_q_class = final_q_class if final_q_class else 'LOAI_1'

        old_is_counted = (old_qc_status in ['PASS', 'PASSNOCHEM']) and (safe_old_q_class == 'LOAI_1')
        new_is_counted = (final_status in ['PASS', 'PASSNOCHEM']) and (safe_final_q_class == 'LOAI_1')
        
        weight_diff = 0
        if not old_is_counted and new_is_counted:
            weight_diff = coil_weight   
        elif old_is_counted and not new_is_counted:
            weight_diff = -coil_weight  
            
        if weight_diff != 0 and order_id:
            cursor.execute("""
                SELECT fulfilled_weight, total_weight 
                FROM order_production_rules WITH (UPDLOCK, ROWLOCK) 
                WHERE [Order] = ?
            """, (order_id,))
            order_info = cursor.fetchone()
            
            if order_info:
                curr_fulfilled = float(order_info[0] or 0)
                total_allowed = float(order_info[1] or 0)
                new_fulfilled = max(0, curr_fulfilled + weight_diff)
                
                cursor.execute("UPDATE order_production_rules SET fulfilled_weight = ? WHERE [Order] = ?", (new_fulfilled, order_id))
                
                if weight_diff > 0: 
                    if prod_status == 'MTO':
                        # NẾU CÒN ROOM: Lấy so_mapping, nếu rỗng thì gán '1' (Chờ SO)
                        if new_fulfilled <= total_allowed:
                            final_mapped_po = so_mapping if so_mapping else '1'
                        else:
                            final_mapped_po = '0'
                    else:
                        final_mapped_po = '0'
                elif weight_diff < 0: 
                    final_mapped_po = '0'
        
        # -----------------------------------------------------------------
        # C. ĐIỀU PHỐI GIA CÔNG
        # -----------------------------------------------------------------
        final_rework_status = old_rework_status
        if not is_downgraded:
            # Gộp chung PASS và PASSNOCHEM
            if final_status in ['PASS', 'PASSNOCHEM']:
                # 🌟 THÊM MỚI: Nếu là PASSNOCHEM và đang LAY_MAU thì GIỮ NGUYÊN để chờ kết quả Cơ tính
                if final_status == 'PASSNOCHEM' and old_rework_status == 'LAY_MAU':
                    final_rework_status = 'LAY_MAU'
                
                # CÁC LỆNH CÒN LẠI HOẶC NẾU ĐÃ PASS: Lên thẳng FINAL / SKIN_CUST
                elif old_rework_status in ['LAY_MAU', 'CXL', 'NULL', None, 'SKIN', 'RCL', 'SKIN_CUST']:
                    if is_skin_required == 1 and not id_xuly:
                        final_rework_status = 'SKIN_CUST' 
                    else:
                        final_rework_status = 'FINAL'
                else:
                    final_rework_status = old_rework_status
                    
            # Gộp chung FAIL và FAILNOCHEM
            else: 
                # (Giữ nguyên logic của phần FAIL/FAILNOCHEM ở hướng dẫn trước)
                if old_rework_status in ['FINAL', 'NULL', '', None]:
                    if is_skin_required == 1 and not id_xuly:
                        final_rework_status = 'SKIN_CUST'
                    else:
                        final_rework_status = None 
                elif actual_qc_stage == 'STAGE_2' and new_stage2_penalty > 0:
                    final_rework_status = old_rework_status 
                else:
                    final_rework_status = old_rework_status 

        # -----------------------------------------------------------------
        # D. LƯU XUỐNG DATABASE
        # -----------------------------------------------------------------
        cursor.execute("""
            UPDATE coil_data 
            SET scores = ?, is_checked = ?, note_qc = ?,
                stage1_penalty = ?, stage1_msg = ?, 
                stage2_penalty = ?, stage2_msg = ?, 
                qc_msg = ?, qc_status = ?, mapped_po = ?,
                quality_class = ?, prime_status = ?, rework_status = ?,
                qc_stage = ?,  
                updated_at = GETDATE()       
            WHERE coil_id = ?
        """, (scores_json_str, new_is_checked, final_note_str, 
            new_stage1_penalty, new_stage1_msg, 
            new_stage2_penalty, new_stage2_msg, 
            final_msg, final_status, final_mapped_po, 
            final_q_class, final_p_status, final_rework_status, 
            actual_qc_stage, coil_id))
        
        msg_to_frontend = f"Lưu thành công. Trạng thái cuộn: {final_status}"
        conn.commit() 
        LAST_DATA_UPDATE['timestamp'] = time.time()
        return jsonify({
            'status': 'success', 
            'msg': msg_to_frontend, 
            'qc_status': final_status,
            'rework_status': final_rework_status,
            'mapped_po': final_mapped_po,
            'qc_msg': final_msg,
            'quality_class': final_q_class, 
            'prime_status': final_p_status,
            'merged_scores': final_scores,
            'merged_notes': final_note_str
        })
        # 🔺 KẾT THÚC KÉO LÙI LỀ

    except Exception as e:
        if conn: conn.rollback() 
        print(f"ERROR save_manual_data: {e}")
        return jsonify({'status':'error', 'msg': str(e)})
    finally:
        if conn: conn.close()


# 1. API KHỞI TẠO BẢNG LOG VÀ CÁC CHỨC NĂNG LIÊN QUAN
@dashboard_bp.route('/api/init_log_table', methods=['GET'])
def init_log_table():
    try:
        conn = db.get_connection()
        # Tạo bảng
        conn.execute("""
            CREATE TABLE IF NOT EXISTS system_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT, message TEXT, log_type TEXT,
                is_read INTEGER DEFAULT 0,
                created_at TEXT
            )
        """)
        try: conn.execute("ALTER TABLE system_logs ADD COLUMN is_read INTEGER DEFAULT 0")
        except: pass

        conn.execute("CREATE INDEX IF NOT EXISTS idx_is_read ON system_logs (is_read)")
        
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'msg': 'DB Logs Optimized!'})
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
@dashboard_bp.route('/api/mark_all_read', methods=['POST'])
def mark_all_read():
    conn = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor() # [SỬA 1]: Tạo cursor
        
        cursor.execute("SELECT TOP 1 1 FROM system_logs WITH (NOLOCK) WHERE is_read = 0")
        check = cursor.fetchone()
        
        if check:
            cursor.execute("UPDATE system_logs SET is_read = 1 WHERE is_read = 0")
            conn.commit()
            
        return jsonify({'status': 'success'})
    except Exception as e:
        print(f"Error mark_all_read: {e}") # In lỗi ra để dễ debug
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close() # Đảm bảo đóng kết nối
# API LOG
@dashboard_bp.route('/api/get_unread_count', methods=['GET'])
def get_unread_count():
    f_id = request.args.get('factory', 'ALL')
    try:
        conn = db.get_connection()
        cursor = conn.cursor() 
        cursor.execute("""
            SELECT COUNT(*) as cnt FROM system_logs WITH (NOLOCK) 
            WHERE is_read = 0 AND (factory = ? OR factory = 'ALL')
        """, (f_id,))
        row = cursor.fetchone()
        conn.close()
        return jsonify({'count': row[0]}) 
    except:
        return jsonify({'count': 0})
# 2. API LẤY LỊCH SỬ THÔNG BÁO
@dashboard_bp.route('/api/get_system_logs', methods=['GET'])
def get_system_logs():
    f_id = request.args.get('factory', 'ALL')
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT TOP 50 * FROM system_logs WITH (NOLOCK) 
            WHERE (factory = ? OR factory = 'ALL') 
            ORDER BY id DESC
        """, (f_id,))
        columns = [column[0] for column in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        conn.close()
        
        logs = []
        for r in rows:
            logs.append({
                'id': r['id'],
                'title': r['title'],
                'message': r['message'],
                'type': r['log_type'],
                'time': r['created_at']
            })
        return jsonify(logs)
    except Exception as e:
        return jsonify([])

# Hàm nội bộ để các file khác gọi (Lưu log vào DB)
def log_system_event(title, message, log_type='info',factory='ALL'):
    conn = None 
    try:
        import datetime
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT TOP 1 message, created_at FROM system_logs WITH (NOLOCK) 
            WHERE factory = ? ORDER BY id DESC
        """, (factory,))
        columns = [column[0] for column in cursor.description]
        last_log_rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

        if last_log_rows:
            last_log = last_log_rows[0]
            if last_log['message'] == message:
                last_time = last_log['created_at']
                if isinstance(last_time, str):
                    try:
                        clean_t = last_time.replace('T', ' ').split('.')[0]
                        last_time = datetime.datetime.strptime(clean_t, "%Y-%m-%d %H:%M:%S")
                    except: pass
                now_time = datetime.datetime.now()
                if isinstance(last_time, datetime.datetime):
                    diff = (now_time - last_time).total_seconds()
                    if diff < 60: 
                        # BỎ CÂU conn.close() Ở ĐÂY VÌ FINALLY SẼ TỰ ĐỘNG CHẠY TRƯỚC KHI RETURN
                        print(f"♻️ [Anti-Spam] Bỏ qua log trùng: {message}")
                        return
                        
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # Đổi thành cursor.execute cho đồng nhất
        cursor.execute(
            "INSERT INTO system_logs (title, message, log_type, is_read, created_at, factory) VALUES (?, ?, ?, 0, ?, ?)",
            (title, message, log_type, now_str, factory)
        )
        conn.commit()
    except Exception as e:
        print(f"❌ Lỗi ghi log: {e}")
    finally:
        if conn: 
            conn.close()
@dashboard_bp.route('/api/get_latest_coils', methods=['GET'])
def get_latest_coils():
    try:
        last_time = request.args.get('since') 
        factory = request.args.get('factory', 'ALL')
        grade = request.args.get('grade', 'ALL')
        selected_ca = request.args.get('Ca', 'ALL')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        raw_coil_ids = request.args.get('coil_ids', '')
        raw_order_ids = request.args.get('order_ids', '')
        db_grade_clean = str(grade).strip().upper()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        sql = """
            SELECT c.coil_id, c.ID_XuLy, c.grade, c.raw_data, c.scores, c.is_checked, 
                c.production_date, c.updated_at, c.factory,
                c.Temperature, c.Speed, c.quality_level, c.note_qc, c.slab_grade, c.weight, c.target_thick, c.target_width, c.Nhom,
                c.qc_stage, c.qc_status, c.mapped_po, c.rework_status,
                c.suggested_order_map, c.qc_msg,
                c.stage1_penalty, c.stage2_penalty,
                c.downgrade_reason,
                ISNULL(r.is_skin_required, 0) as is_skin_required,
                c.quality_class, c.prime_status,
                c.[Order] as original_order,
                tm.tdc_code,
                c.TARGET_LV2,
                tv.criteria_json
            FROM coil_data c WITH (NOLOCK) 
            LEFT JOIN order_production_rules r WITH (NOLOCK) ON c.[Order] = r.[Order]
            LEFT JOIN tdc_versions tv WITH (NOLOCK) ON c.target_tdc_version_id = tv.id
            LEFT JOIN tdc_master tm WITH (NOLOCK) ON tv.master_id = tm.id
            WHERE (c.updated_at > ? OR c.production_date > ?) 
        """
        params = [last_time, last_time]
        
        coil_id_list = []
        order_id_list = []
        if raw_coil_ids:
            raw_list = re.split(r'[\s,;]+', raw_coil_ids)
            coil_id_list = [x.strip().upper() for x in raw_list if x.strip()]
            coil_id_list = list(dict.fromkeys(coil_id_list))
        if raw_order_ids: 
            raw_list_ord = re.split(r'[\s,;]+', raw_order_ids)
            order_id_list = list(dict.fromkeys([x.strip().upper() for x in raw_list_ord if x.strip()]))    
        if coil_id_list:
            if len(coil_id_list) > 1000:
                coil_id_list = coil_id_list[:1000]
            placeholders = ', '.join(['?'] * len(coil_id_list))
            sql += f" AND (c.coil_id IN ({placeholders}) OR c.ID_XuLy IN ({placeholders}))"
            params.extend(coil_id_list)
            params.extend(coil_id_list)
        elif order_id_list:
            if len(order_id_list) > 1000: order_id_list = order_id_list[:1000]
            placeholders = ', '.join(['?'] * len(order_id_list))
            sql += f" AND c.[Order] IN ({placeholders})"
            params.extend(order_id_list)
        else:
            if factory != 'ALL':
                sql += " AND c.factory = ?"
                params.append(factory)
            if grade != 'ALL':
                sql += " AND c.grade = ?"
                params.append(grade)
            if selected_ca != 'ALL':
                sql += " AND c.Ca = ?"
                params.append(selected_ca)
            if start_date:
                s_str = start_date.replace('T', ' ')
                if len(s_str) == 10: s_str += " 00:00:00"
                elif len(s_str) == 16: s_str += ":00"
                sql += " AND c.production_date >= ?"
                params.append(s_str)
            if end_date:
                e_str = end_date.replace('T', ' ')
                if len(e_str) == 10: e_str += " 23:59:59.999"
                elif len(e_str) == 16: e_str += ":59.999"
                sql += " AND c.production_date <= ?"
                params.append(e_str)
            
        cursor.execute(sql, tuple(params))
        rows = db.fetchall_as_dict(cursor)
        conn.close()
        
        # Lấy config để tính lại auto_scores (nét đứt)
        all_configs = get_all_grade_configs()
        
        new_data = {}
        for r in rows:
            cid = r['coil_id']
            raw = json.loads(r['raw_data']) if r['raw_data'] else {}
            current_scores = json.loads(r['scores']) if r['scores'] else {}
            grade_clean = (r['grade'] or 'SAE1006').strip().upper()

            # 1. Tính auto_scores (Dùng cho Nét đứt)
            if raw:
                thick_val = float(r.get('TARGET_LV2') or 0.0)
                auto_scores = process_coil_scores(cid, raw, grade_clean, thickness=thick_val, cached_config=all_configs)
            else:
                auto_scores = {}
            raw_crit = r.get('criteria_json')
            tdc_limits = {}
            if raw_crit:
                try:
                    crit_list = json.loads(raw_crit) if isinstance(raw_crit, str) else raw_crit
                    for crit in crit_list:
                        defect = crit.get('defect')
                        allowed_range = crit.get('range', [])
                        if defect and allowed_range:
                            tdc_limits[defect] = {
                            'min': min(allowed_range),
                            'max': max(allowed_range)
                        }
                except: pass
            # 2. Đưa điểm ra lớp ngoài cùng (Dùng cho Nét liền)
            frontend_obj = current_scores.copy() 
            frontend_obj['auto_scores'] = auto_scores
            frontend_obj['tdc_limits'] = tdc_limits
            # 3. Tối ưu chuỗi hiển thị khi Hover vào Radar
            optimized_raw = {}
            if raw:
                for k, v in raw.items():
                    # --- ĐOẠN NÀY LÀ CODE MỚI CẦN THÊM VÀO ---
                    grade_config = all_configs.get(db_grade_clean, all_configs.get('SAE1006'))

                    if isinstance(v, list):
                        cfg = grade_config.get(k, {})
                        optimized_raw[k] = build_matrix_tooltip(v, cfg)
                    else:
                        optimized_raw[k] = v

            frontend_obj['raw_data'] = optimized_raw
            frontend_obj['GRADE'] = grade_clean
            frontend_obj['IS_CHECKED'] = r['is_checked']
            frontend_obj['updated_at'] = r['updated_at']
            frontend_obj['production_date'] = str(r['production_date']) if r['production_date'] else ''
            frontend_obj['Temperature'] = r['Temperature'] or 0
            frontend_obj['Speed'] = r['Speed'] or 0
            frontend_obj['quality_level'] = r['quality_level'] or ''
            frontend_obj['note_qc'] = r['note_qc'] or ''
            frontend_obj['slab_grade'] = r['slab_grade'] or '---'
            frontend_obj['weight'] = float(r['weight'] or 0.0)
            frontend_obj['target_thick'] = float(r['target_thick'] or 0)
            frontend_obj['target_width'] = float(r['target_width'] or 0)
            frontend_obj['ID_xuly'] = r['ID_XuLy'] or ''
            frontend_obj['Nhom'] = r['Nhom'] or ''
            frontend_obj['qc_stage'] = r['qc_stage'] or ''
            frontend_obj['qc_status'] = r['qc_status'] or ''
            frontend_obj['mapped_po'] = r['mapped_po'] or ''
            frontend_obj['rework_status'] = r['rework_status'] or ''
            frontend_obj['suggested_order_map'] = r['suggested_order_map'] or ''
            frontend_obj['is_skin_required'] = r['is_skin_required']
            frontend_obj['qc_msg'] = r['qc_msg'] or ''
            frontend_obj['quality_class'] = r.get('quality_class', '')  
            frontend_obj['prime_status'] = r.get('prime_status', '')
            frontend_obj['downgrade_reason'] = r.get('downgrade_reason', '')
            frontend_obj['original_order'] = r['original_order'] or '---'
            frontend_obj['tdc_code'] = r['tdc_code'] or '---'
            frontend_obj['stage1_penalty'] = r.get('stage1_penalty', 0) 
            frontend_obj['stage2_penalty'] = r.get('stage2_penalty', 0)
            new_data[cid] = frontend_obj
        
        return jsonify({'status': 'success', 'data': new_data})
    except Exception as e:
        print(f"Lỗi API get_latest_coils: {e}")
        return jsonify({'status': 'error', 'msg': str(e)})
@dashboard_bp.route('/api/import_notes_excel', methods=['POST'])
def import_notes_excel():
    print("\n" + "="*50)
    print("🚀 [START] Bắt đầu gọi API import_notes_excel")
    start_total = time.time()

    if 'file' not in request.files:
        return jsonify({'status': 'error', 'msg': 'Không tìm thấy file.'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'status': 'error', 'msg': 'Chưa chọn file.'})
    
    try:
        # ---------------------------------------------------------
        # CHỐT 1: ĐO THỜI GIAN ĐỌC EXCEL
        # ---------------------------------------------------------
        print("⏳ [1/4] Đang đọc file Excel (pd.read_excel)...")
        t0 = time.time()
        df = pd.read_excel(file)
        t1 = time.time()
        print(f"✅ [1/4] Xong! Đọc Excel mất: {t1 - t0:.3f} giây. Số dòng thô: {len(df)}")
        
        if len(df.columns) < 2:
            return jsonify({'status': 'error', 'msg': 'File Excel phải có ít nhất 2 cột: Cột 1 là ID, Cột 2 là Note.'})
        
        # ---------------------------------------------------------
        # CHỐT 2: ĐO THỜI GIAN XỬ LÝ DATAFRAME
        # ---------------------------------------------------------
        print("⏳ [2/4] Đang làm sạch dữ liệu...")
        t2 = time.time()
        id_col = df.columns[0]
        note_col = df.columns[1]
        
        df = df.dropna(subset=[id_col])
        df[id_col] = df[id_col].astype(str).str.strip().str.upper()
        df[note_col] = df[note_col].fillna('').astype(str).str.strip()
        df = df.drop_duplicates(subset=[id_col], keep='last')

        update_params = []
        updated_dict = {}
        for val_id, val_note in zip(df[id_col], df[note_col]):
            if not val_id or val_id == 'NAN': 
                continue
            update_params.append((val_note, val_id))
            updated_dict[val_id] = val_note
            
        t3 = time.time()
        print(f"✅ [2/4] Xong! Xử lý DF mất: {t3 - t2:.3f} giây. Số dòng hợp lệ cần Update: {len(update_params)}")

        if not update_params:
            return jsonify({'status': 'error', 'msg': 'Không có dữ liệu hợp lệ trong file.'})

        # ---------------------------------------------------------
        # CHỐT 3: ĐO THỜI GIAN KẾT NỐI DB & EXECUTE TỪNG CHUNK
        # ---------------------------------------------------------
        print("⏳ [3/4] Đang kết nối Database...")
        t4 = time.time()
        conn = None
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.fast_executemany = True
            print(f"✅ Kết nối DB mất: {time.time() - t4:.3f} giây.")
            
            sql_by_coil = """
                UPDATE coil_data 
                SET note_qc = ?, updated_at = GETDATE()
                WHERE coil_id = ?
            """
            
            sql_by_idxuly = """
                UPDATE coil_data 
                SET note_qc = ?, updated_at = GETDATE()
                WHERE ID_XuLy = ?
            """
            
            print("⏳ [4/4] BẮT ĐẦU UPDATE XUỐNG DB...")
            t_db_start = time.time()
            chunk_size = 500 
            
            for i in range(0, len(update_params), chunk_size):
                chunk = update_params[i:i + chunk_size]
                print(f"   -> Đang chạy Chunk từ {i} đến {i + len(chunk)}...")
                
                # Đo thời gian cho coil_id
                t_coil_1 = time.time()
                cursor.executemany(sql_by_coil, chunk)
                print(f"      + Xong coil_id mất: {time.time() - t_coil_1:.3f} giây")
                
                # Đo thời gian cho ID_XuLy
                t_idxuly_1 = time.time()
                cursor.executemany(sql_by_idxuly, chunk)
                print(f"      + Xong ID_XuLy mất: {time.time() - t_idxuly_1:.3f} giây")
                
            print("⏳ Đang Commit Transaction...")
            conn.commit()
            print(f"✅ [4/4] Update & Commit xong! Chạy Database mất: {time.time() - t_db_start:.3f} giây")
            
        except Exception as db_err:
            if conn:
                conn.rollback() 
            raise db_err 
            
        finally:
            if conn:
                conn.close() 

        t_end_total = time.time()
        print(f"🎉 [FINISH] TỔNG THỜI GIAN API HOẠT ĐỘNG: {t_end_total - start_total:.3f} giây")
        print("="*50 + "\n")

        return jsonify({
            'status': 'success', 
            'msg': f'Đã xử lý file và cập nhật Ghi chú thành công cho {len(update_params)} mã.',
            'updated_data': updated_dict
        })

    except Exception as e:
        import traceback
        print(f"❌ LỖI RỒI: {traceback.format_exc()}")
        return jsonify({'status': 'error', 'msg': f'Lỗi đọc file: {str(e)}'})
# =========================================================================
# GIAI ĐOẠN 3: CÁC API XỬ LÝ NGOẠI LỆ & ĐIỀU PHỐI (REWORK & DOWNGRADE)
# =========================================================================

@dashboard_bp.route('/api/update_rework_status', methods=['POST'])
@permission_required('qlcl_input')
def update_rework_status():
    conn = None
    try:
        req = request.json
        coil_id = req.get('coil_id')
        rework_type = req.get('rework_type') 
        user = session.get('username', 'Unknown')

        if not coil_id:
            return jsonify({'status': 'error', 'msg': 'Thiếu thông tin ID cuộn thép'})

        conn = db.get_connection()
        cursor = conn.cursor()

        import datetime
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # --- VÁ LỖI SỐ 2: KHÓA DÒNG VÀ KIỂM TRA TRẠNG THÁI TRƯỚC KHI UPDATE ---
        cursor.execute("""
            SELECT qc_status, rework_status 
            FROM coil_data WITH (UPDLOCK, ROWLOCK)
            WHERE coil_id = ?
        """, (coil_id,))
        
        row = cursor.fetchone()
        if not row:
            return jsonify({'status': 'error', 'msg': 'Không tìm thấy cuộn thép.'})
            
        current_qc_status, current_rework_status = row
        
        # CHẶN MÙ QUÁNG: Nếu cuộn này đã bị người khác chốt Hạ cấp/Phế (FINAL) hoặc PASS
        if current_qc_status == 'PASS' or (current_rework_status == 'FINAL' and current_qc_status not in ['PASSNOCHEM', 'FAILNOCHEM', 'PENDING']):
            return jsonify({
                'status': 'error', 
                'msg': f'Thao tác thất bại! Cuộn {coil_id} đã được người khác chốt phân cấp hoặc đã ĐẠT. Không thể phát lệnh {rework_type}.'
            })

        # --- BẮT ĐẦU CẬP NHẬT (KÈM VÁ LỖI CẬP NHẬT THỜI GIAN updated_at) ---
        if rework_type in ['CANNOT_REWORK', 'LAY_MAU']:
            cursor.execute("""
                UPDATE coil_data 
                SET rework_status = ?, updated_at = GETDATE()
                WHERE coil_id = ?
            """, (rework_type, coil_id))
        else:
            # Nếu đi xử lý (SKIN, MÀI), Xóa ID_xuly cũ để chờ xưởng cập nhật mã mới
            cursor.execute("""
                UPDATE coil_data 
                SET ID_xuly = NULL, rework_status = ?, updated_at = GETDATE()
                WHERE coil_id = ?
            """, (rework_type, coil_id))

        # Ghi Log lịch sử
        cursor.execute("""
            INSERT INTO audit_log_qlcl (coil_id, user_name, defect_key, old_value, new_value, changed_at) 
            VALUES (?, ?, ?, ?, ?, ?)
        """, (coil_id, user, f'REWORK_ACTION: {rework_type}', 0, 1, now_str))

        conn.commit()
        return jsonify({'status': 'success', 'msg': f'Đã phát lệnh điều phối hình thức [{rework_type}] lên hệ thống thành công!'})

    except Exception as e:
        if conn: conn.rollback() # Trả lại Lock nếu có lỗi
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
@dashboard_bp.route('/api/downgrade_coil', methods=['POST'])
@permission_required('qlcl_input')
def downgrade_coil():
    req = request.json
    coil_id = req.get('coil_id')
    action_type = req.get('action_type')  # 'NON_PRIME' hoặc 'SCRAP'
    note = req.get('note', '')
    user = session.get('username', 'Unknown')

    if not coil_id:
        return jsonify({'status': 'error', 'msg': 'Thiếu thông tin ID cuộn thép'})

    conn = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # 1. Lấy dữ liệu hiện tại
        cursor.execute("""
            SELECT c.weight, c.[Order], r.production_status, r.SO_mapping, c.note_qc, c.mapped_po, c.qc_status, c.quality_class
            FROM coil_data c 
            LEFT JOIN order_production_rules r ON c.[Order] = r.[Order]
            WHERE c.coil_id = ?
        """, (coil_id,))
        
        row = cursor.fetchone() 
        if not row:
            return jsonify({'status': 'error', 'msg': 'Không tìm thấy cuộn thép.'})
            
        (coil_weight, order_id, prod_status, so_mapping, old_note_raw, old_mapped_po, old_qc_status, old_q_class) = row
        coil_weight = float(coil_weight) if coil_weight is not None else 0.0
        if order_id is not None:
            order_id = str(order_id).strip()
        # 2. Logic tính toán và cập nhật sản lượng đơn hàng (🌟 ĐÃ FIX LỖI TRỪ KÉP & BẢO TOÀN NON-PRIME)
        if order_id:
            cursor.execute("SELECT fulfilled_weight FROM order_production_rules WITH (UPDLOCK, ROWLOCK) WHERE [Order] = ?", (order_id,))
            order_row = cursor.fetchone()
            if order_row:
                current_fulfilled = float(order_row[0] or 0)
                
                # QUY TẮC VÀNG: Cuộn này trước đó có đang được tính vào sản lượng không? (Phải PASS và là LOẠI 1)
                is_currently_counted = (old_qc_status == 'PASS') and (old_q_class == 'LOAI_1')
                
                # SỬA LẠI LOGIC NÀY:
                new_fulfilled = current_fulfilled
                should_update_order = False

                if is_currently_counted and action_type == 'SCRAP':
                    # Đang tính mà bị vứt đi -> Trừ ra
                    new_fulfilled = max(0, current_fulfilled - coil_weight)
                    should_update_order = True
                elif not is_currently_counted and action_type == 'NON_PRIME':
                    # Đang rớt mà được hạ cấp thành LOAI_1 -> Cộng vào
                    new_fulfilled = current_fulfilled + coil_weight
                    should_update_order = True
                    
                if should_update_order:
                    cursor.execute("""
                        UPDATE order_production_rules 
                        SET fulfilled_weight = ? 
                        WHERE [Order] = ?
                    """, (new_fulfilled, order_id))
        # 3. XỬ LÝ LÝ DO HẠ CẤP (Ghi trực tiếp vào cột mới, TUYỆT ĐỐI KHÔNG ĐỤNG note_qc)
        import datetime
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Tạo chuỗi ghi chú rõ ràng có kèm tên người hạ cấp
        action_log = f"[{now_str} - {user}]: {note}"

        # 4. Định nghĩa cấp chất lượng
        final_q_class = 'LOAI_1' if action_type == 'NON_PRIME' else 'LOAI_2'
        final_p_status = action_type 

        # 5. CẬP NHẬT CHỐT: Ghi vào downgrade_reason
        new_mapped_po = '0' # Mặc định là Tồn kho (MTS)
        if action_type == 'NON_PRIME':
            if prod_status == 'MTO':
                if so_mapping:
                    # Nếu đã có SO mapping từ Kế hoạch -> Kéo sang
                    new_mapped_po = so_mapping
                else:
                    # Nếu là hàng Đặt (MTO) nhưng chưa có mã SO -> Gán cờ '1' (Chờ SO)
                    new_mapped_po = '1'
        
        # 5. CẬP NHẬT CHỐT: Ép mapped_po vào câu SQL
        if action_type == 'SCRAP':
            # Phế thì luôn trả về 0
            cursor.execute("""
                UPDATE coil_data 
                SET prime_status = ?, quality_class = ?, 
                    downgrade_reason = ?, mapped_po = '0', rework_status = 'FINAL', qc_status = 'PASS'
                WHERE coil_id = ?
            """, (final_p_status, final_q_class, action_log, coil_id))
        else:
            # NON-PRIME thì dùng biến new_mapped_po vừa tính toán
            cursor.execute("""
                UPDATE coil_data 
                SET prime_status = ?, quality_class = ?, 
                    downgrade_reason = ?, mapped_po = ?, rework_status = 'FINAL', qc_status = 'PASS'
                WHERE coil_id = ?
            """, (final_p_status, final_q_class, action_log, new_mapped_po, coil_id))

        # 6. Ghi log lịch sử thay đổi
        cursor.execute("""
            INSERT INTO audit_log_qlcl (coil_id, user_name, defect_key, old_value, new_value, changed_at) 
            VALUES (?, ?, ?, ?, ?, ?)
        """, (coil_id, user, f'DOWNGRADE: {action_type}', 0, 1, now_str))

        conn.commit()
        
        # 🌟 TRẢ VỀ new_mapped_po LÊN UI ĐỂ ĐỒNG BỘ RAM
        return jsonify({
            'status': 'success', 
            'msg': f'Hệ thống đã phê duyệt hạ cấp sang {action_type} và khóa cuộn thành công!',
            'downgrade_reason': action_log,
            'mapped_po': '0' if action_type == 'SCRAP' else new_mapped_po
        })

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
@dashboard_bp.route('/api/recommend_order', methods=['POST'])
def recommend_order():
    conn = None
    try:
        req = request.json
        coil_id = req.get('coil_id')

        if not coil_id:
            return jsonify({'status': 'error', 'msg': 'Thiếu ID cuộn thép'})

        conn = db.get_connection()
        cursor = conn.cursor()

        # 1. Lấy thông tin cuộn hiện tại
        cursor.execute("""
            SELECT c.weight, c.scores, c.qc_stage, r.KySanXuat, r.[material_desc], LEFT(r.[Order], 1) as factory_prefix
            FROM coil_data c WITH (NOLOCK)
            LEFT JOIN order_production_rules r WITH (NOLOCK) ON c.[Order] = r.[Order]
            WHERE c.coil_id = ?
        """, (coil_id,))
        coil_info = cursor.fetchone()
        if not coil_info:
            return jsonify({'status': 'error', 'msg': 'Không tìm thấy cuộn.'})

        coil_weight, scores_str, qc_stage, ky_sx, mat_desc, factory_prefix = coil_info
        coil_weight = float(coil_weight) if coil_weight else 0.0
        scores_json = json.loads(scores_str) if scores_str else {}

        # 2. Quét thô danh sách Order (BỎ điều kiện fulfilled_weight, LẤY thêm req_min_w, req_max_w)
        cursor.execute("""
            SELECT r.[Order], r.target_tdc_version_id, v.criteria_json, r.production_status, r.SO_mapping, 
                   ISNULL(r.fulfilled_weight, 0), ISNULL(r.total_weight, 0), 
                   ISNULL(r.req_min_w, 0), ISNULL(r.req_max_w, 0)
            FROM order_production_rules r WITH (NOLOCK)
            JOIN tdc_versions v WITH (NOLOCK) ON r.target_tdc_version_id = v.id
            WHERE r.KySanXuat = ? 
              AND r.[material_desc] = ?
              AND LEFT(r.[Order], 1) = ?
        """, (ky_sx, mat_desc, factory_prefix))
        
        candidates = cursor.fetchall()
        suggested_orders = []

        from utils.scoring import evaluate_tdc_stage_1, evaluate_tdc_stage_2

        # 3. Giả lập chạy qua từng tiêu chuẩn
        for cand in candidates:
            (order_id, tdc_id, criteria_json, prod_status, so_mapping, 
             fulfilled_w, total_w, req_min_w, req_max_w) = cand
             
            fulfilled_w, total_w = float(fulfilled_w), float(total_w)
            
            # --- Chạy mô phỏng Stage 1 (Truyền đúng min/max khối lượng) ---
            res1 = evaluate_tdc_stage_1(scores_json, criteria_json, coil_weight, req_min_w, req_max_w)
            total_penalty = res1['stage1_penalty']
            
            # --- Chạy mô phỏng Stage 2 (Nếu cuộn đã có cơ tính) ---
            if qc_stage == 'STAGE_2':
                res2 = evaluate_tdc_stage_2(scores_json, criteria_json)
                total_penalty += res2['stage2_penalty']

            # Nếu mô phỏng PASS -> Đưa vào danh sách ứng viên
            if total_penalty == 0:
                # Đánh giá Room (Còn dung lượng hay không)
                has_room = (fulfilled_w + coil_weight) <= total_w
                remaining = total_w - fulfilled_w
                
                suggested_orders.append({
                    'order_id': order_id,
                    'so_mapping': so_mapping,
                    'prod_status': prod_status,
                    'total_weight': total_w,
                    'fulfilled_weight': fulfilled_w,
                    'remaining_weight': remaining,
                    'has_room': has_room,
                    'req_min': req_min_w,
                    'req_max': req_max_w
                })

        # 4. Sắp xếp ưu tiên (Ranking Logic)
        # Tiêu chí 1: has_room = True xếp trước (Nhóm ưu tiên)
        # Tiêu chí 2: Cùng nhóm thì đơn nào có tổng total_weight lớn hơn (hoặc remaining lớn hơn) xếp trước
        suggested_orders.sort(key=lambda x: (not x['has_room'], -x['remaining_weight']))

        return jsonify({'status': 'success', 'data': suggested_orders})

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
from sqlalchemy import text
@dashboard_bp.route('/api/apply_order', methods=['POST'])
@permission_required('qlcl_input')
def apply_order():
    data = request.json
    coil_id = data.get('coil_id')
    order_id = data.get('order_id')
    
    if not coil_id or not order_id:
        return jsonify({'status': 'error', 'msg': 'Thiếu tham số'}), 400

    try:
        with db.engine.begin() as conn: 
            conn.execute(text("""
                UPDATE coil_data 
                SET 
                    suggested_order_map = :oid,
                    -- Cập nhật thông báo để UI hiển thị trạng thái chờ SAP
                    qc_msg = N'⏳ ĐANG CHỜ SAP ĐỔI SANG ĐƠN: ' + :oid
                WHERE coil_id = :cid
            """), {"oid": order_id, "cid": coil_id})
            
        return jsonify({
            'status': 'success', 
            'msg': f'Đã ghi nhận gợi ý chờ SAP chuyển sang đơn {order_id}.'
        })
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)}), 500
@dashboard_bp.route('/api/remove_mapped_po', methods=['POST'])
@permission_required('qlcl_remove_po')
def remove_mapped_po():
    conn = None
    try:
        req = request.json
        coil_id = req.get('coil_id')
        user = session.get('username', 'Unknown') # Lấy tên user thao tác

        if not coil_id:
            return jsonify({'status': 'error', 'msg': 'Thiếu ID cuộn thép'})

        conn = db.get_connection()
        cursor = conn.cursor()

        # 1. Lấy thông tin mapped_po cũ để ghi log
        cursor.execute("SELECT mapped_po FROM coil_data WITH (NOLOCK) WHERE coil_id = ?", (coil_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'status': 'error', 'msg': 'Không tìm thấy cuộn.'})
            
        old_mapped_po = row[0] or '0'

        # 2. Update mapped_po về '0' (Tồn kho)
        cursor.execute("""
            UPDATE coil_data 
            SET mapped_po = '0', updated_at = GETDATE()
            WHERE coil_id = ?
        """, (coil_id,))

        # 3. Ghi Log lịch sử
        import datetime
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("""
            INSERT INTO audit_log_qlcl (coil_id, user_name, defect_key, old_value, new_value, changed_at) 
            VALUES (?, ?, ?, ?, ?, ?)
        """, (coil_id, user, 'REMOVE_MAPPED_PO', str(old_mapped_po), '0', now_str))

        conn.commit()
        return jsonify({'status': 'success', 'msg': f'Đã gỡ cờ Mapped PO thành công! Cuộn {coil_id} đã về trạng thái Tồn kho.'})

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
@dashboard_bp.route('/api/complete_cxl', methods=['POST'])
@permission_required('qlcl_input')
def complete_cxl():
    conn = None
    try:
        req = request.json
        coil_id = req.get('coil_id')
        user = session.get('username', 'Unknown')

        if not coil_id:
            return jsonify({'status': 'error', 'msg': 'Thiếu ID cuộn thép'})

        conn = db.get_connection()
        cursor = conn.cursor()

        import datetime
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 1. TÌM TRẠNG THÁI HIỆN TẠI TRƯỚC KHI XÓA
        cursor.execute("SELECT rework_status FROM coil_data with (nolock) WHERE coil_id = ?", (coil_id,))
        row = cursor.fetchone()
        
        # Nếu cuộn không tồn tại hoặc trạng thái hiện tại không hợp lệ
        valid_statuses = ['CXL', 'RCL', 'SKIN', 'SKIN_CUST', 'LAY_MAU']
        if not row or row[0] not in valid_statuses:
            return jsonify({'status': 'error', 'msg': 'Cuộn không ở trạng thái gia công hợp lệ để mở khóa.'})
            
        current_rework = row[0] # Lưu lại tên trạng thái (VD: 'SKIN_CUST', 'LAY_MAU')

        # 2. ĐƯA TRẠNG THÁI VỀ NULL (Như code gốc của bạn)
        cursor.execute("""
            UPDATE coil_data 
            SET rework_status = 'NULL', updated_at = GETDATE()
            WHERE coil_id = ?
        """, (coil_id,))

        # 3. GHI LOG CHÍNH XÁC TÊN TRẠNG THÁI VỪA MỞ KHÓA
        cursor.execute("""
            INSERT INTO audit_log_qlcl (coil_id, user_name, defect_key, old_value, new_value, changed_at) 
            VALUES (?, ?, ?, ?, ?, ?)
        """, (coil_id, user, f'UNLOCK_{current_rework}', 0, 1, now_str))

        conn.commit()
        return jsonify({'status': 'success', 'msg': f'Đã mở khóa cuộn từ trạng thái {current_rework} để đánh giá lại!'})

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
@dashboard_bp.route('/api/undo_downgrade', methods=['POST'])
@permission_required('qlcl_manager_undo') # Đảm bảo quyền này đã được cấu hình trong DB phân quyền
def undo_downgrade():
    req = request.json
    coil_id = req.get('coil_id')
    user = session.get('username', 'Unknown')

    if not coil_id:
        return jsonify({'status': 'error', 'msg': 'Thiếu ID cuộn thép'})

    conn = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # 1. Khóa dòng và lấy trạng thái cuộn hiện tại
        cursor.execute("""
            SELECT c.weight, c.[Order], c.prime_status 
            FROM coil_data c WITH (UPDLOCK, ROWLOCK)
            WHERE c.coil_id = ?
        """, (coil_id,))
        row = cursor.fetchone()

        if not row:
            return jsonify({'status': 'error', 'msg': 'Không tìm thấy cuộn thép.'})

        coil_weight, order_id, p_status = row
        coil_weight = float(coil_weight) if coil_weight else 0.0
        if order_id is not None:
            order_id = str(order_id).strip()
        if p_status not in ['NON_PRIME', 'SCRAP']:
            return jsonify({'status': 'error', 'msg': 'Cuộn này không ở trạng thái hạ cấp để hoàn tác.'})

        # 2. Xử lý trả lại Sản lượng (Order Weight)
        if order_id:
            cursor.execute("SELECT fulfilled_weight FROM order_production_rules WITH (UPDLOCK, ROWLOCK) WHERE [Order] = ?", (order_id,))
            order_row = cursor.fetchone()
            if order_row:
                current_fulfilled = float(order_row[0] or 0)
                if p_status == 'NON_PRIME':
                    new_fulfilled = max(0, current_fulfilled - coil_weight)
                    cursor.execute("""
                        UPDATE order_production_rules 
                        SET fulfilled_weight = ? 
                        WHERE [Order] = ?
                    """, (new_fulfilled, order_id))

        # 3. Reset các cột về trạng thái chờ
        import datetime
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cursor.execute("""
            UPDATE coil_data 
            SET prime_status = NULL, 
                quality_class = NULL, 
                downgrade_reason = NULL, 
                mapped_po = NULL, 
                rework_status = 'NULL', 
                qc_status = 'FAIL',
                updated_at = GETDATE()
            WHERE coil_id = ?
        """, (coil_id,))

        # 4. Ghi log lịch sử hệ thống
        log_action_name = f'UNDO_DOWNGRADE: {p_status} -> FAIL'
        cursor.execute("""
            INSERT INTO audit_log_qlcl (coil_id, user_name, defect_key, old_value, new_value, changed_at) 
            VALUES (?, ?, ?, ?, ?, ?)
        """, (coil_id, user, log_action_name, 1, 0, now_str))

        conn.commit()
        return jsonify({'status': 'success', 'msg': 'Đã hủy lệnh hạ cấp! Cuộn được trả về trạng thái chờ đánh giá.'})

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        if conn: conn.close()
@dashboard_bp.route('/api/get_preview_tdc', methods=['POST'])
@permission_required('qlcl_input')
def get_preview_tdc():
    """API Lấy khung tiêu chuẩn TDC dựa vào mã Order để ướm thử lên Radar"""
    req = request.json
    order_id = req.get('order_id')
    if not order_id:
        return jsonify({'status': 'error', 'msg': 'Thiếu mã Order'})

    conn = db.get_connection()
    cursor = conn.cursor()
    try:
        # [SỬA Ở ĐÂY]: JOIN thêm tdc_master để lấy tdc_code
        cursor.execute("""
            SELECT v.criteria_json, tm.tdc_code 
            FROM order_production_rules r WITH (NOLOCK)
            JOIN tdc_versions v WITH (NOLOCK) ON r.target_tdc_version_id = v.id
            JOIN tdc_master tm WITH (NOLOCK) ON v.master_id = tm.id
            WHERE r.[Order] = ?
        """, (order_id,))
        row = cursor.fetchone()
        
        if not row or not row[0]:
            return jsonify({'status': 'error', 'msg': 'Không tìm thấy Order hoặc Order này chưa được gán tiêu chuẩn TDC!'})
        
        criteria_list = json.loads(row[0])
        tdc_code = row[1] if row[1] else '---' # Lấy mã TDC
        
        tdc_limits = {}
        for crit in criteria_list:
            defect = crit.get('defect')
            allowed_range = crit.get('range', [])
            if defect and allowed_range:
                tdc_limits[defect] = {
                    'min': min(allowed_range),
                    'max': max(allowed_range)
                }
                
        # Trả về thêm tdc_code cho Frontend
        return jsonify({'status': 'success', 'tdc_limits': tdc_limits, 'tdc_code': tdc_code})
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        conn.close()
@dashboard_bp.route('/api/get_preview_order_list', methods=['GET'])
@permission_required('qlcl_input')
def get_preview_order_list():
    """API lấy danh sách Order cho Dropdown tìm kiếm nhanh"""
    conn = db.get_connection()
    cursor = conn.cursor()
    try:
        # Lấy danh sách Order kèm mô tả sản phẩm để KCS dễ tìm
        cursor.execute("""
            SELECT DISTINCT r.[Order], r.material_desc, r.KySanXuat 
            FROM order_production_rules r WITH (NOLOCK)
            WHERE r.target_tdc_version_id IS NOT NULL
            ORDER BY r.KySanXuat DESC, r.[Order] DESC
        """)
        rows = cursor.fetchall()
        
        orders = []
        for r in rows:
            if r[0]:
                orders.append({
                    'order_id': r[0],
                    'desc': f"{r[0]} | {r[1]}" # Ghép mã Order và Mô tả để dễ search
                })
                
        return jsonify({'status': 'success', 'data': orders})
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)})
    finally:
        conn.close()